#!/usr/bin/env python3
"""
oob_monitor.py

Cong cu giam sat, Verify vat ly va Tu dong phuc hoi (Self-healing) menu OOB.
Kien truc Multi-Terminal:
    - Terminal 1: Giam sat lien tuc (Daemon) + Deep Verify (Smart Clear).
    - Terminal 2: Menu Quan ly - Them/Xoa IP, Cau hinh, Xem danh sach, Log, Manual Push.
"""

import getpass
import json
import os
import platform
import sqlite3
import subprocess
import sys
import time
import re
import threading
import concurrent.futures
from datetime import datetime, timedelta
from collections import deque

from rich import box as rbox
from rich.console import Console, Group
from rich.panel import Panel
from rich.rule import Rule
from rich.table import Table
from rich.text import Text
from rich.layout import Layout
from rich.live import Live

# Import tu oob_lib
from oob_lib import (
    poll_host, MiniTelnet, connect_auto, fetch_hostname,
    fetch_hostname_via_auto, hostname_matches_description,
    push_menu_descriptions, ping_host
)

CONFIG_FILE_DEFAULT = "oob_config.json"

DEFAULT_CONFIG = {
    "username": "",
    "password": "",
    "enable_password": "",
    "vertiv_connect_password": "",  
    "credentials": [],          
    "menu_name_override": "",
    "ssh_port": 22,
    "telnet_port": 23,
    "interval": 30,
    "verify_interval": 3600,
    "ip_list": "oob_ips.txt",
    "baseline_db": "baseline.db",
    "snapshot_db": "snapshot.db",
    "auto_verify": True,        
    "verify_schedule_mode": "interval", 
    "verify_schedule_time": "01:00",    
    "verify_schedule_weekday": "mon",   
    "scan_schedule_mode": "interval",   
    "scan_schedule_time": "01:00",      
    "scan_schedule_weekday": "mon",     
    "verify_wait_after_connect": 1.5,
    "max_verify_duration": 300,
}

# ---------------------------------------------------------------------------
# Locks de dong bo luong
# ---------------------------------------------------------------------------
db_lock = threading.Lock()            
action_lock = threading.Lock()        
file_lock = threading.Lock()          
ui_print_lock = threading.Lock()      

# ---------------------------------------------------------------------------
# Regex
# ---------------------------------------------------------------------------
_DESC_PREFIX_RE    = re.compile(r'^[-=>\s]+')
_HOSTNAME_PROMPT_RE = re.compile(r'([A-Za-z0-9_\-\.]+)[>#]')
_HOSTNAME_LOGIN_RE  = re.compile(r'([A-Za-z0-9_\-\.]+)\s+login:', re.IGNORECASE)
_HOSTNAME_BSD_RE    = re.compile(r'(?:\()?([A-Za-z0-9_\-\.]+)(?:\))?\s*\(tty', re.IGNORECASE)
_ANSI_STRIP_RE      = re.compile(r'\x1b\[.*?m')
_CONN_ERR_RE        = re.compile(r'refused|time(d)?[\s-]?out|unreachable|no route to host|unknown host|% ', re.IGNORECASE)
_IP_RE              = re.compile(r'^\d{1,3}(\.\d{1,3}){3}$')

# ---------------------------------------------------------------------------
# He thong UI da luong
# ---------------------------------------------------------------------------
_con = Console(highlight=False)

MAX_LOG = 15
oob_logs = deque(maxlen=MAX_LOG)
verify_logs = deque(maxlen=MAX_LOG)
ui_lock = threading.Lock()

layout = Layout()
layout.split_column(
    Layout(name="upper"),
    Layout(name="lower")
)

_live_ui = None

def update_ui():
    with ui_lock:
        layout["upper"].update(Panel(Text.from_markup("\n".join(oob_logs)), title="[bold cyan] [ OOB MONITORING (Cau hinh) ] [/]", border_style="cyan"))
        layout["lower"].update(Panel(Text.from_markup("\n".join(verify_logs)), title="[bold magenta] [ DEEP VERIFY (Thiet bi cuoi) ] [/]", border_style="magenta"))
        if _live_ui and _live_ui.is_started:
            _live_ui.update(layout)

def log_oob(msg: str):
    ts = datetime.now().strftime("%H:%M:%S")
    oob_logs.append(f"[dim]\\[{ts}][/] {msg}")
    update_ui()

def log_verify(msg: str):
    ts = datetime.now().strftime("%H:%M:%S")
    verify_logs.append(f"[dim]\\[{ts}][/] {msg}")
    update_ui()

def log_baseline_change(alias, ip, action):
    os.makedirs("baseline-logs", exist_ok=True)
    log_path = os.path.join("baseline-logs", "baseline_updates.log")
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        with file_lock:
            with open(log_path, "a", encoding="utf-8") as f:
                f.write(f"[{ts}] {action}: {alias} ({ip})\n")
    except Exception: pass

# ---------------------------------------------------------------------------
# Trang thai thiet bi
# ---------------------------------------------------------------------------
DEVICE_STATUS_FILE = "device_status.json"
_device_status_lock = threading.Lock()

def load_device_status() -> dict:
    if not os.path.exists(DEVICE_STATUS_FILE): return {}
    try:
        with open(DEVICE_STATUS_FILE, "r", encoding="utf-8") as f: return json.load(f)
    except (json.JSONDecodeError, OSError): return {}

def save_device_status(ip, **fields):
    with _device_status_lock:
        data = load_device_status()
        entry = data.get(ip, {})
        entry.update(fields)
        entry["checked_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        data[ip] = entry
        try:
            with open(DEVICE_STATUS_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
        except OSError: pass

MENU_STATE_LABELS = {
    "ok": "OK",
    "no_menu": "Khong co menu",
    "fetch_failed": "Khong lay duoc thong tin menu",
}

# ---------------------------------------------------------------------------
# Cac ham tien ich, cau hinh va DB
# ---------------------------------------------------------------------------
def load_config(path):
    cfg = DEFAULT_CONFIG.copy()
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f: cfg.update(json.load(f))
        except (json.JSONDecodeError, OSError) as exc: pass
    return cfg

def save_config(path, cfg):
    with open(path, "w", encoding="utf-8") as f: json.dump(cfg, f, indent=2, ensure_ascii=False)

def mask(value): return "******" if value else "(chua dat)"

import base64
try:
    from cryptography.fernet import Fernet
    _FERNET_AVAIL = True
except ImportError:
    _FERNET_AVAIL = False

def _get_cipher():
    key_file = ".oob_secret.key"
    if _FERNET_AVAIL:
        if not os.path.exists(key_file):
            with open(key_file, "wb") as f:
                f.write(Fernet.generate_key())
        with open(key_file, "rb") as f:
            return Fernet(f.read())
    return None

def _encrypt_cred(text):
    if not text: return text
    c = _get_cipher()
    if c:
        return "ENC:" + c.encrypt(text.encode("utf-8")).decode("utf-8")
    return "B64:" + base64.b64encode(text.encode("utf-8")).decode("utf-8")

def _decrypt_cred(text):
    if not text: return text
    if text.startswith("ENC:"):
        c = _get_cipher()
        if c:
            try: return c.decrypt(text[4:].encode("utf-8")).decode("utf-8")
            except Exception: pass
        return text
    elif text.startswith("B64:"):
        try: return base64.b64decode(text[4:].encode("utf-8")).decode("utf-8")
        except Exception: pass
    return text

def _encrypt_dict(d):
    return {k: _encrypt_cred(v) if k in ["password", "enable_password"] else v for k, v in d.items()}

def _decrypt_dict(d):
    return {k: _decrypt_cred(v) if k in ["password", "enable_password"] else v for k, v in d.items()}

def save_working_credential(ip, cred):
    if not ip or not cred: return
    cache = {}
    try:
        if os.path.exists("working_creds.json"):
            with open("working_creds.json", "r", encoding="utf-8") as f:
                cache = json.load(f)
    except Exception: pass
    
    enc_cred = _encrypt_dict(cred)
    old_cred = cache.get(ip)
    if old_cred:
        old_dec = _decrypt_dict(old_cred)
        if old_dec == cred: return

    cache[ip] = enc_cred
    try:
        with open("working_creds.json", "w", encoding="utf-8") as f:
            json.dump(cache, f, indent=2)
    except Exception: pass

def get_all_credentials(cfg, oob_ip=None):
    creds = []
    working_cred = None
    if oob_ip:
        try:
            if os.path.exists("working_creds.json"):
                with open("working_creds.json", "r", encoding="utf-8") as f:
                    cache = json.load(f)
                    if oob_ip in cache:
                        working_cred = _decrypt_dict(cache[oob_ip])
        except Exception: pass
    
    if working_cred: creds.append(working_cred)

    if cfg.get("username") or cfg.get("password"):
        default_cred = {"username": cfg.get("username", ""), "password": cfg.get("password", ""), "enable_password": cfg.get("enable_password", "")}
        if default_cred not in creds: creds.append(default_cred)
        
    for c in cfg.get("credentials", []):
        if c not in creds: creds.append(c)
        
    if not creds: creds.append({"username": "", "password": "", "enable_password": ""})
    return creds

_WEEKDAY_MAP = {"mon": 0, "tue": 1, "wed": 2, "thu": 3, "fri": 4, "sat": 5, "sun": 6}
_WEEKDAY_LABELS = {"mon": "Thu 2", "tue": "Thu 3", "wed": "Thu 4", "thu": "Thu 5", "fri": "Thu 6", "sat": "Thu 7", "sun": "Chu nhat"}

def _parse_hhmm(text, default="01:00"):
    try:
        h_str, m_str = str(text).strip().split(":")
        h, m = int(h_str), int(m_str)
        if 0 <= h <= 23 and 0 <= m <= 59: return h, m
    except (ValueError, AttributeError): pass
    dh, dm = default.split(":")
    return int(dh), int(dm)

def compute_next_scheduled_run(mode, time_str, weekday_str, now=None):
    now = now or datetime.now()
    hh, mm = _parse_hhmm(time_str)
    candidate = now.replace(hour=hh, minute=mm, second=0, microsecond=0)
    if mode == "weekly":
        target_wd = _WEEKDAY_MAP.get((weekday_str or "mon").strip().lower()[:3], 0)
        days_ahead = (target_wd - now.weekday()) % 7
        candidate = candidate + timedelta(days=days_ahead)
        if candidate <= now: candidate += timedelta(days=7)
        return candidate
    if candidate <= now: candidate += timedelta(days=1)
    return candidate

def _describe_verify_schedule(cfg):
    mode = cfg.get("verify_schedule_mode", "interval")
    if mode == "daily": return f"Hang ngay luc {cfg.get('verify_schedule_time', '01:00')}"
    if mode == "weekly": return f"Hang tuan vao {_WEEKDAY_LABELS.get((cfg.get('verify_schedule_weekday', 'mon') or 'mon').strip().lower()[:3], 'mon')} luc {cfg.get('verify_schedule_time', '01:00')}"
    return f"Lap lai moi {cfg.get('verify_interval', 3600)}s"

def _describe_scan_schedule(cfg):
    mode = cfg.get("scan_schedule_mode", "interval")
    if mode == "daily": return f"Hang ngay luc {cfg.get('scan_schedule_time', '01:00')}"
    if mode == "weekly": return f"Hang tuan vao {_WEEKDAY_LABELS.get((cfg.get('scan_schedule_weekday', 'mon') or 'mon').strip().lower()[:3], 'mon')} luc {cfg.get('scan_schedule_time', '01:00')}"
    return f"Lap lai moi {cfg.get('interval', 30)}s"

def _edit_verify_schedule(cfg):
    _con.print(f"\n  --- Lich chay Deep Verify tu dong ---")
    _con.print(f"  Hien tai: {_describe_verify_schedule(cfg)}")
    _con.print("  1. Lap lai theo chu ky (interval)")
    _con.print("  2. Hang ngay, vao 1 gio co dinh")
    _con.print("  3. Hang tuan, vao 1 thu + gio co dinh")
    mode_choice = _con.input("  [cyan]Chon che do (1/2/3)[/]: ").strip()
    if mode_choice == "1":
        cfg["verify_schedule_mode"] = "interval"
        _con.print("  [green](OK)[/] Da chuyen ve che do lap lai theo chu ky.")
    elif mode_choice == "2":
        cfg["verify_schedule_mode"] = "daily"
        val = _con.input(f"  [cyan]Gio chay moi ngay (HH:MM)[/]: ").strip()
        if val:
            h, m = _parse_hhmm(val)
            cfg["verify_schedule_time"] = f"{h:02d}:{m:02d}"
        _con.print(f"  [green](OK)[/] Da dat lich: Hang ngay luc {cfg.get('verify_schedule_time', '01:00')}.")
    elif mode_choice == "3":
        cfg["verify_schedule_mode"] = "weekly"
        wd_val = _con.input(f"  [cyan]Thu (mon/tue/wed/thu/fri/sat/sun)[/]: ").strip().lower()
        if wd_val[:3] in _WEEKDAY_MAP: cfg["verify_schedule_weekday"] = wd_val[:3]
        val = _con.input(f"  [cyan]Gio chay (HH:MM)[/]: ").strip()
        if val:
            h, m = _parse_hhmm(val)
            cfg["verify_schedule_time"] = f"{h:02d}:{m:02d}"
        _con.print(f"  [green](OK)[/] Da dat lich: {_describe_verify_schedule(cfg)}.")
    else:
        _con.print("  [yellow][!][/] Lua chon khong hop le.")

def _edit_scan_schedule(cfg):
    _con.print(f"\n  --- Lich chay Thu thap cau hinh ---")
    _con.print(f"  Hien tai: {_describe_scan_schedule(cfg)}")
    _con.print("  1. Lap lai theo chu ky (interval)")
    _con.print("  2. Hang ngay, vao 1 gio co dinh")
    _con.print("  3. Hang tuan, vao 1 thu + gio co dinh")
    mode_choice = _con.input("  [cyan]Chon che do (1/2/3)[/]: ").strip()
    if mode_choice == "1":
        cfg["scan_schedule_mode"] = "interval"
        _con.print("  [green](OK)[/] Da chuyen ve che do lap lai theo chu ky.")
    elif mode_choice == "2":
        cfg["scan_schedule_mode"] = "daily"
        val = _con.input(f"  [cyan]Gio chay moi ngay (HH:MM)[/]: ").strip()
        if val:
            h, m = _parse_hhmm(val)
            cfg["scan_schedule_time"] = f"{h:02d}:{m:02d}"
        _con.print(f"  [green](OK)[/] Da dat lich: Hang ngay luc {cfg.get('scan_schedule_time', '01:00')}.")
    elif mode_choice == "3":
        cfg["scan_schedule_mode"] = "weekly"
        wd_val = _con.input(f"  [cyan]Thu (mon/tue/wed/thu/fri/sat/sun)[/]: ").strip().lower()
        if wd_val[:3] in _WEEKDAY_MAP: cfg["scan_schedule_weekday"] = wd_val[:3]
        val = _con.input(f"  [cyan]Gio chay (HH:MM)[/]: ").strip()
        if val:
            h, m = _parse_hhmm(val)
            cfg["scan_schedule_time"] = f"{h:02d}:{m:02d}"
        _con.print(f"  [green](OK)[/] Da dat lich: {_describe_scan_schedule(cfg)}.")
    else:
        _con.print("  [yellow][!][/] Lua chon khong hop le.")

def settings_menu(cfg, config_path):
    while True:
        v_note = "[dim red]<- Khong hieu luc[/]" if cfg.get("verify_schedule_mode", "interval") in ("daily", "weekly") else "[dim green]<- Dang co hieu luc[/]"
        s_note = "[dim red]<- Khong hieu luc[/]" if cfg.get("scan_schedule_mode", "interval") in ("daily", "weekly") else "[dim green]<- Dang co hieu luc[/]"
        auto_v = "[green bold]BAT[/]" if cfg.get('auto_verify', True) else "[red bold]TAT[/]"

        g = Table.grid(padding=(0, 1))
        g.add_column(style="bold cyan", min_width=4, justify="right"); g.add_column()
        g.add_row("", "[dim]-- KET NOI ----------------------------------------------------------[/]")
        g.add_row("[1]",  f"Username (chinh)      : {cfg['username'] or '[dim](khong dung)[/]'}")
        g.add_row("[2]",  f"Password (chinh)      : {mask(cfg['password'])}")
        g.add_row("[3]",  f"Enable pass (chinh)   : {mask(cfg['enable_password'])}")
        g.add_row("\\[y]", f"Vertiv Connect Pass   : {mask(cfg.get('vertiv_connect_password', ''))}")
        g.add_row("\\[k]", f"Tai khoan phu (multi) : [bold]{len(cfg.get('credentials', []))} tai khoan[/]")
        g.add_row("[5]",  f"SSH port (uu tien)    : {cfg.get('ssh_port', 22)}")
        g.add_row("[6]",  f"Telnet port (du phong): {cfg.get('telnet_port', 23)}")
        g.add_row("", "")
        g.add_row("", "[dim]-- FILE DU LIEU ----------------------------------------------------[/]")
        g.add_row("[8]",  f"File danh sach IP     : {cfg['ip_list']}")
        g.add_row("[9]",  f"File baseline DB      : {cfg['baseline_db']}")
        g.add_row("\\[a]",  f"File snapshot DB      : {cfg['snapshot_db']}")
        g.add_row("", "")
        g.add_row("", "[dim]-- LUONG 1: GIAM SAT CAU HINH ------------------------------------[/]")
        g.add_row("[4]",  f"Ten menu ep dung      : {cfg['menu_name_override'] or '[dim](tu dong do)[/]'}")
        g.add_row("\\[s]", f"Lich chay Thu thap    : [cyan]{_describe_scan_schedule(cfg)}[/]")
        g.add_row("[7]",  f"Chu ky interval (s)   : [bold cyan]{cfg['interval']}[/]  {s_note}")
        g.add_row("", "")
        g.add_row("", "[dim]-- LUONG 2: VERIFY VAT LY ----------------------------------------[/]")
        g.add_row("\\[b]", f"Tu dong Verify ngam   : {auto_v}")
        g.add_row("\\[d]", f"Lich chay Verify      : [cyan]{_describe_verify_schedule(cfg)}[/]")
        g.add_row("\\[v]", f"Chu ky interval (s)   : [bold cyan]{cfg.get('verify_interval', 3600)}[/]  {v_note}")
        g.add_row("\\[w]", f"Cho sau connect (s)   : [bold cyan]{cfg.get('verify_wait_after_connect', 1.5)}[/]")
        g.add_row("\\[m]", f"Timeout Verify (s)    : [bold cyan]{cfg.get('max_verify_duration', 300)}[/]")
        g.add_row("", "")
        g.add_row("\\[t]",  "[bold yellow]Thu ket noi nhanh (test credential)[/]")
        g.add_row("[0]",   "[bold red]Quay lai menu chinh[/]")

        _con.print()
        _con.print(Panel(g, title="[bold cyan] [ CAI DAT HE THONG ] [/]", border_style="cyan", padding=(1, 2)))
        choice = _con.input("[bold]Chon muc can sua[/]: ").strip().lower()

        if choice == "1": cfg["username"] = input("  Username moi (Enter de bo trong): ").strip()
        elif choice == "2": cfg["password"] = getpass.getpass("  Password moi: ").strip()
        elif choice == "3": cfg["enable_password"] = getpass.getpass("  Enable password moi: ").strip()
        elif choice == "y": cfg["vertiv_connect_password"] = getpass.getpass("  Vertiv Connect Pass: ").strip()
        elif choice == "4": cfg["menu_name_override"] = input(f"  Ten menu ep dung: ").strip()
        elif choice == "5": val = input("  SSH port moi: ").strip(); cfg["ssh_port"] = int(val) if val.isdigit() else cfg["ssh_port"]
        elif choice == "6": val = input("  Telnet port moi: ").strip(); cfg["telnet_port"] = int(val) if val.isdigit() else cfg["telnet_port"]
        elif choice == "7": val = input("  Chu ky thu thap (giay): ").strip(); cfg["interval"] = int(val) if val.isdigit() else cfg["interval"]
        elif choice == "v": val = input("  Chu ky Verify vat ly (giay): ").strip(); cfg["verify_interval"] = int(val) if val.isdigit() else cfg["verify_interval"]
        elif choice == "w": val = input("  Cho sau connect (giay): ").strip(); cfg["verify_wait_after_connect"] = round(float(val), 2) if val else cfg["verify_wait_after_connect"]
        elif choice == "m": val = input("  Timeout tong Verify (giay): ").strip(); cfg["max_verify_duration"] = int(val) if val.isdigit() and int(val)>=30 else cfg["max_verify_duration"]
        elif choice == "8": val = input("  File danh sach IP moi: ").strip(); cfg["ip_list"] = val if val else cfg["ip_list"]
        elif choice == "9": val = input("  File baseline DB moi: ").strip(); cfg["baseline_db"] = val if val else cfg["baseline_db"]
        elif choice == "a": val = input("  File snapshot DB moi: ").strip(); cfg["snapshot_db"] = val if val else cfg["snapshot_db"]
        elif choice == "b": cfg["auto_verify"] = not cfg.get("auto_verify", True)
        elif choice == "d": _edit_verify_schedule(cfg)
        elif choice == "s": _edit_scan_schedule(cfg)
        elif choice == "k": 
            _con.print("\n  [bold]Danh sach Tai khoan phu (Multi-Account)[/]")
            creds = cfg.setdefault("credentials", [])
            if not creds: _con.print("  [dim](Chua co tai khoan phu nao)[/]")
            for i, c in enumerate(creds):
                _con.print(f"  {i+1}. User: [cyan]{c.get('username')}[/] | Pass: ***")
            _con.print("\n  [a] Them tai khoan moi  |  [d] Xoa 1 tai khoan (theo so)  |  [c] Xoa tat ca  |  [0] Quay lai")
            sub = input("  Chon: ").strip().lower()
            if sub == "a":
                u = input("  Username: ").strip()
                p = getpass.getpass("  Password: ").strip()
                e = getpass.getpass("  Enable Password: ").strip()
                creds.append({"username": u, "password": p, "enable_password": e})
                _con.print("  [green](OK)[/] Da them.")
            elif sub == "d":
                if not creds: _con.print("  [yellow][!][/] Danh sach tai khoan phu dang trong.")
                else:
                    idx_raw = input(f"  Nhap so thu tu can xoa (1-{len(creds)}): ").strip()
                    if idx_raw.isdigit() and 1 <= int(idx_raw) <= len(creds):
                        removed = creds.pop(int(idx_raw) - 1)
                        _con.print(f"  [green](OK)[/] Da xoa tai khoan: {removed.get('username')}")
            elif sub == "c":
                cfg["credentials"] = []
                _con.print("  [green](OK)[/] Da xoa toan bo tai khoan phu.")
        elif choice == "t":
            test_ip = _con.input("  [cyan]Nhap IP OOB can thu ket noi[/]: ").strip()
            if test_ip:
                creds = get_all_credentials(cfg, test_ip)
                _con.print(f"  [cyan][*][/] Dang thu ket noi toi [bold]{test_ip}[/] ...")
                last_exc = None
                for idx, c in enumerate(creds):
                    try:
                        session = connect_auto(test_ip, cfg.get("ssh_port", 22), cfg.get("telnet_port", 23), c["username"], c["password"], c["enable_password"], timeout=8)
                        hn = fetch_hostname(session)
                        session.close()
                        save_working_credential(test_ip, c)
                        _con.print(f"  [green](OK)[/] Ket noi thanh cong ({c['username']})! Hostname: [bold cyan]{hn or '?'}[/]")
                        break
                    except Exception as e: last_exc = e
                else:
                    _con.print(f"  [red][!][/] Ket noi that bai: {last_exc}")
        elif choice == "0":
            save_config(config_path, cfg)
            _con.print(f"[green][*][/] Da luu cau hinh vao {config_path}")
            return
        else:
            _con.print("[yellow][!][/] Lua chon khong hop le.")
            continue
        save_config(config_path, cfg)

def load_ip_list(path):
    hosts = []
    try:
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                if line.strip() and not line.startswith("#"):
                    parts = line.split()
                    hosts.append((parts[0], parts[1] if len(parts) > 1 else parts[0]))
    except FileNotFoundError: pass
    return hosts

_ip_list_cache: dict = {"path": None, "mtime": 0, "hosts": []}
def load_ip_list_cached(path: str) -> list:
    try: mtime = os.path.getmtime(path)
    except OSError: return []
    if _ip_list_cache["path"] == path and _ip_list_cache["mtime"] == mtime: return _ip_list_cache["hosts"]
    hosts = load_ip_list(path)
    _ip_list_cache.update({"path": path, "mtime": mtime, "hosts": hosts})
    return hosts

def add_ip(path, ip, alias=None):
    hosts = load_ip_list(path)
    if any(h[0] == ip for h in hosts): return _con.print(f"  [yellow][!][/] IP {ip} da co.")
    with open(path, "a", encoding="utf-8") as f: f.write(f"{ip} {alias or ip}\n")
    _con.print(f"  [green](OK)[/] Da them {ip}")

def remove_ip(path, ip):
    hosts = load_ip_list(path)
    remaining = [h for h in hosts if h[0] != ip]
    if len(remaining) == len(hosts): return _con.print(f"  [yellow][!][/] Khong tim thay IP.")
    with open(path, "w", encoding="utf-8") as f:
        for h_ip, alias in remaining: f.write(f"{h_ip} {alias}\n")
    _con.print(f"  [green](OK)[/] Da xoa {ip}")

_DB_INIT_CACHE = set()
def _init_db(path, table):
    conn = sqlite3.connect(path)
    if (path, table) in _DB_INIT_CACHE: return conn
    conn.execute(f"CREATE TABLE IF NOT EXISTS {table} (host TEXT NOT NULL, menu_name TEXT NOT NULL, option_key TEXT NOT NULL, device_name TEXT, description TEXT, target_ip TEXT NOT NULL, target_port INTEGER NOT NULL, updated_at TEXT NOT NULL, PRIMARY KEY (host, menu_name, option_key))")
    cols = [row[1] for row in conn.execute(f"PRAGMA table_info({table})")]
    if "device_name" not in cols: conn.execute(f"ALTER TABLE {table} ADD COLUMN device_name TEXT")
    if "protocol" not in cols: conn.execute(f"ALTER TABLE {table} ADD COLUMN protocol TEXT DEFAULT 'telnet'")
    if "raw_key" not in cols: conn.execute(f"ALTER TABLE {table} ADD COLUMN raw_key TEXT")
    if "real_menu_name" not in cols: conn.execute(f"ALTER TABLE {table} ADD COLUMN real_menu_name TEXT")
    if "vendor" not in cols: conn.execute(f"ALTER TABLE {table} ADD COLUMN vendor TEXT DEFAULT 'cisco'")
    conn.commit()
    _DB_INIT_CACHE.add((path, table))
    return conn

def get_options_by_host(db_path, table, host):
    with db_lock:
        conn = _init_db(db_path, table)
        rows = conn.execute(f"SELECT menu_name, option_key, device_name, description, target_ip, target_port, protocol, raw_key, real_menu_name, vendor FROM {table} WHERE host=?", (host,)).fetchall()
        conn.close()
    if not rows: return None, None, None
    menu_name, device_name, options = rows[0][0], rows[0][2], {}
    for _mn, key, _dn, desc, ip, port, proto, raw_key, real_menu_name, vendor in rows:
        options[key] = {
            "description": desc, "ip": ip, "port": port, "protocol": proto, 
            "_raw_key": raw_key if raw_key else key, "_menu_name": real_menu_name if real_menu_name else _mn,
            "vendor": vendor if vendor else "cisco"
        }
    return menu_name, device_name, options

def get_updated_at_by_host(db_path, table, host):
    with db_lock:
        conn = _init_db(db_path, table)
        row = conn.execute(f"SELECT MAX(updated_at) FROM {table} WHERE host=?", (host,)).fetchone()
        conn.close()
    return row[0] if row and row[0] else None

def save_options(db_path, table, host, menu_name, device_name, options):
    now = datetime.now().isoformat(timespec="seconds")
    with db_lock:
        conn = _init_db(db_path, table)
        conn.execute(f"DELETE FROM {table} WHERE host=? AND menu_name=?", (host, menu_name))
        for key, entry in options.items():
            conn.execute(
                f"INSERT INTO {table} (host, menu_name, option_key, device_name, description, target_ip, target_port, protocol, raw_key, real_menu_name, vendor, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (host, menu_name, key, device_name, entry.get("description", ""), entry["ip"], entry.get("port", 23), entry.get("protocol", "telnet"), entry.get("_raw_key", key), entry.get("_menu_name", menu_name), entry.get("vendor", "cisco"), now)
            )
        conn.commit()
        conn.close()

def _norm_proto(p): return p or "telnet"

def options_equal(a: dict, b: dict) -> bool:
    if set(a.keys()) != set(b.keys()): return False
    for key in a:
        if a[key].get("ip") != b[key].get("ip") or a[key].get("port") != b[key].get("port") or _norm_proto(a[key].get("protocol")) != _norm_proto(b[key].get("protocol")) or a[key].get("description") != b[key].get("description"): return False
    return True

def diff_options(baseline: dict, snapshot: dict) -> dict:
    extra, missing, changed = sorted(set(snapshot) - set(baseline)), sorted(set(baseline) - set(snapshot)), []
    for key in sorted(set(baseline) & set(snapshot)):
        b, s = baseline[key], snapshot[key]
        if b.get("ip") != s.get("ip") or b.get("port") != s.get("port") or _norm_proto(b.get("protocol")) != _norm_proto(s.get("protocol")) or b.get("description") != s.get("description"): changed.append(key)
    return {"extra": extra, "missing": missing, "changed": changed}

def _fmt_entry(e: dict) -> str: return f"{_norm_proto(e.get('protocol'))}://{e.get('ip','')}:{e.get('port', 22 if _norm_proto(e.get('protocol')) == 'ssh' else 23)}"

def print_diff(baseline: dict, snapshot: dict):
    d = diff_options(baseline, snapshot)
    if d["extra"]: _con.print(f"    [red]+ Option la (them): {', '.join(d['extra'])}[/]")
    if d["missing"]: _con.print(f"    [red]- Option bi mat: {', '.join(d['missing'])}[/]")
    if d["changed"]:
        _con.print(f"    [yellow]~ Option bi doi noi dung: {', '.join(d['changed'])}[/]")
        t = Table(box=rbox.ASCII, header_style="bold dim", show_header=True)
        t.add_column("Option",  style="bold cyan"); t.add_column("Baseline (chuan)"); t.add_column("Hien tai")
        for key in d["changed"]: t.add_row(f"{key}", f"{baseline[key].get('description','')} [dim]{_fmt_entry(baseline[key])}[/]", f"{snapshot[key].get('description','')} [dim]{_fmt_entry(snapshot[key])}[/]")
        _con.print(t)

def print_options(options: dict):
    t = Table(box=rbox.ASCII, show_header=False, padding=(0, 1))
    t.add_column(style="bold cyan", justify="right"); t.add_column(); t.add_column(style="dim")
    for key in sorted(options):
        e, proto = options[key], _norm_proto(options[key].get("protocol"))
        t.add_row(f"{key}", e.get("description", ""), f"[{'green' if proto == 'ssh' else 'yellow'}]{proto}[/]://{e['ip']}:{e.get('port', 22 if proto == 'ssh' else 23)}")
    _con.print(t)

# ---------------------------------------------------------------------------
# Thu thap Menu (Cisco & Vertiv)
# ---------------------------------------------------------------------------
MENU_NAME_RE = re.compile(r'^\s*menu\s+(\S+)\s+(?:text|command)\b', re.IGNORECASE | re.MULTILINE)
TEXT_RE       = re.compile(r'^\s*menu\s+(\S+)\s+text\s+(\S+)\s+(.+)', re.IGNORECASE)
CMD_TELNET_RE = re.compile(r'^\s*menu\s+(\S+)\s+command\s+(\S+)\s+telnet\s+(\S+)(?:\s+(\d+))?', re.IGNORECASE)
CMD_SSH_RE    = re.compile(r'^\s*menu\s+(\S+)\s+command\s+(\S+)\s+ssh\s+(?:-l\s+\S+\s+|\S+@)?(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})(?:\s+(\d+))?', re.IGNORECASE)

def poll_host_multi(ip, cfg, timeout=10):
    menu_name_override = cfg.get("menu_name_override") or None
    creds = get_all_credentials(cfg, ip)
    last_exc = None
    for c in creds:
        try:
            tn = connect_auto(ip, cfg.get("ssh_port", 22), cfg.get("telnet_port", 23), c["username"], c["password"], c["enable_password"], timeout=timeout)
            try:
                tn.write("\r\n")
                initial_prompt = tn.read_until(["#", ">", "cli->"], timeout=5)
                save_working_credential(ip, c)

                if "cli->" in initial_prompt:
                    tn.write("cd access/")
                    tn.read_until("cli->", timeout=5)
                    tn.write("show")
                    raw = tn.read_until("cli->", timeout=15)

                    if tn.last_read_timed_out and not raw.strip(): return None, None, {}, "fetch_failed"

                    final_options = {}
                    hostname = None
                    lines = raw.splitlines()
                    for i, line in enumerate(lines):
                        line = line.strip()
                        if line.startswith("===") and i + 1 < len(lines):
                            potential_host = lines[i+1].strip()
                            if potential_host and " " not in potential_host: hostname = potential_host

                        m = re.match(r'^(\S+)\s+(\d+)\s+serial\s+', line, re.IGNORECASE)
                        if m:
                            desc, port_str = m.group(1), m.group(2)
                            key = str(port_str)
                            final_options[key] = {
                                "description": desc, "ip": ip, "port": int(port_str),
                                "protocol": "serial", "_raw_key": key, "_menu_name": "access", "vendor": "vertiv"
                            }

                    if not final_options: return hostname, None, {}, "no_menu"
                    return hostname, "Vertiv ACS", final_options, "ok"

                else:
                    hostname = fetch_hostname(tn)
                    tn.write("terminal length 0"); tn.read_until("#", timeout=5)
                    tn.write("show running-config | include menu")
                    raw = tn.read_until_prompt(timeout=15)
                    if tn.last_read_timed_out and not raw.strip(): return hostname, None, {}, "fetch_failed"

                    detected_names = list(set(MENU_NAME_RE.findall(raw)))
                    if menu_name_override:
                        if menu_name_override not in detected_names: raise ValueError(f"Ten menu ep dung khong co tren thiet bi.")
                        menu_names = [menu_name_override]
                    else: menu_names = detected_names

                    if not menu_names: return hostname, None, {}, "fetch_failed" if tn.last_read_timed_out else "no_menu"
                        
                    all_texts, all_commands = {}, {}
                    for r_line in raw.splitlines():
                        line = r_line.strip()
                        m_t = TEXT_RE.match(line)
                        if m_t and m_t.group(1) in menu_names: all_texts[(m_t.group(1), m_t.group(2).strip())] = m_t.group(3).strip(); continue
                        m_c = CMD_TELNET_RE.match(line)
                        if m_c and m_c.group(1) in menu_names: all_commands[(m_c.group(1), m_c.group(2).strip())] = {"ip": m_c.group(3), "port": int(m_c.group(4)) if m_c.group(4) else 23, "protocol": "telnet"}; continue
                        m_s = CMD_SSH_RE.match(line)
                        if m_s and m_s.group(1) in menu_names: all_commands[(m_s.group(1), m_s.group(2).strip())] = {"ip": m_s.group(3), "port": int(m_s.group(4)) if m_s.group(4) else 22, "protocol": "ssh"}

                    final_options = {}
                    for (m_name, cmd_k), cmd_data in all_commands.items():
                        if (m_name, f"[{cmd_k}]") in all_texts: real_key, desc = f"[{cmd_k}]", all_texts[(m_name, f"[{cmd_k}]")]
                        elif (m_name, cmd_k) in all_texts: real_key, desc = cmd_k, all_texts[(m_name, cmd_k)]
                        else: real_key, desc = cmd_k, ""
                        display_key = real_key[1:-1] if real_key.startswith("[") and real_key.endswith("]") else real_key
                        final_options[f"{m_name} [{display_key}]" if len(menu_names) > 1 else display_key] = {"description": desc, "ip": cmd_data["ip"], "port": cmd_data["port"], "protocol": cmd_data["protocol"], "_raw_key": real_key, "_menu_name": m_name, "vendor": "cisco"}
                    return hostname, " + ".join(sorted(menu_names)), final_options, "ok"
            finally:
                try: tn.write("exit")
                except OSError: pass
                tn.close()
        except Exception as e: last_exc = e
    raise last_exc or RuntimeError("Khong co credential nao hop le hoac danh sach credential trong.")

# ---------------------------------------------------------------------------
# Deep Verify & Alarm
# ---------------------------------------------------------------------------
def _build_verify_report(alias, oob_ip, own_hostname, results):
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    counts = {}
    for r in results: counts[r["status"]] = counts.get(r["status"], 0) + 1
    summary = "   ".join(f"{s}: {counts.get(s, 0)}" for s in ("CANH BAO", "KHONG PIVOT", "TIMEOUT", "YEU CAU DANG NHAP", "OK") if counts.get(s, 0)) or "Khong co opt"
    
    def fmt_row(cols, widths): return "| " + " | ".join(cols[i].ljust(widths[i]) for i in range(len(cols))) + " |"
    def fmt_sep(widths): return "+" + "+".join("-" * (w + 2) for w in widths) + "+"
    
    headers = ["STT", "Option", "Trang thai", "Hostname thuc te", "Description", "Port", "Ghi chu"]
    max_w   = [4, 22, 12, 22, 32, 6, 36]
    ordered = sorted(enumerate(results), key=lambda p: ({"CANH BAO": 0, "KHONG PIVOT": 1, "TIMEOUT": 2, "YEU CAU DANG NHAP": 3, "OK": 4}.get(p[1]["status"], 9), p[0]))
    rows = []
    for i, (_, r) in enumerate(ordered, start=1):
        rows.append([str(i), (r["key"][:max_w[1]-1]+"~") if len(r["key"])>max_w[1] else r["key"], r["status"], (r.get("act_host") or "-")[:max_w[3]], (r.get("desc") or "-")[:max_w[4]], str(r.get("port", "")), (r.get("note") or "")[:max_w[6]]])
    widths = [min(max([len(h)] + ([len(row[i]) for row in rows] if rows else [])), max_w[i]) for i, h in enumerate(headers)]
    
    bar = "=" * max(sum(widths) + 3*len(widths) + 1, 60)
    lines = [bar, f" BAO CAO DEEP VERIFY - OOB: {alias} ({oob_ip})", f" Thoi gian      : {now_str}", f" Hostname OOB   : {own_hostname or '?'}", f" Tong so option : {len(results)}", f" Tom tat        : {summary}", bar]
    lines.extend([fmt_sep(widths), fmt_row(headers, widths), fmt_sep(widths)])
    if rows:
        for row in rows: lines.append(fmt_row(row, widths))
    else: lines.append("| (khong co option nao)".ljust(len(bar)-1) + "|")
    lines.extend([fmt_sep(widths), bar])
    return "\n".join(lines)

def extract_hostname(output: str) -> str:
    """Ham quet va boc tach Hostname tu luong text dau ra cua thiet bi.
    Da duoc nang cap de d?n sach ky tu an (Control Chars) va tang cuong Regex cho FreeBSD."""
    output = _ANSI_STRIP_RE.sub('', output)
    # Loai bo toan bo ky tu dien khien ngoai tru xuong dong (\r, \n)
    output = re.sub(r'[\x00-\x09\x0b\x0c\x0e-\x1f\x7f]', '', output)
    
    auth_seen = False
    lines = [l.strip() for l in output.splitlines() if l.strip()]
    
    for line in reversed(lines):
        if any(x in line for x in ["telnet ", "ssh ", "Trying ", "Open", "Connection refused", "disconnect", "clear line", "Type the hot key", "cli->", "Welcome to ACS"]) or _CONN_ERR_RE.search(line): 
            continue
        
        # 1. Kieu login chuan Linux/Unix: "hostname login:"
        m_login = re.search(r'([A-Za-z0-9_\-\.]+)\s+login:', line, re.IGNORECASE)
        if m_login: return m_login.group(1)
        
        # 2. Kieu FreeBSD (Dung tren nhieu OOB/Firewall): "FreeBSD/amd64 (hostname) (ttyu0)" hoac Linux
        m_bsd = re.search(r'(?:FreeBSD|Linux|NetBSD|OpenBSD).*?\(([A-Za-z0-9_\-\.]+)\)', line, re.IGNORECASE)
        if m_bsd: return m_bsd.group(1)
        
        # 3. Kieu Cisco Console: "hostname con0 is now available"
        m_con = re.search(r'([A-Za-z0-9_\-\.]+)\s+con\d+\s+is now available', line, re.IGNORECASE)
        if m_con: return m_con.group(1)
        
        # 4. Prompt chung: "hostname>" hoac "hostname#"
        m_prompt = _HOSTNAME_PROMPT_RE.search(line)
        if m_prompt: 
            h = m_prompt.group(1)
            # Loai tru cac prompt he thong
            if h.lower() not in ["cli", "cli-", "access", "admin", "root"]: 
                return h
                
        if re.search(r'Username:|Password:|login:', line, re.IGNORECASE): 
            auth_seen = True
            
    return "AUTH_REQUIRED" if auth_seen else None

def run_deep_verify(cfg, alias, oob_ip, options, print_fn=None):
    if print_fn is None: print_fn = log_verify
    max_duration = float(cfg.get("max_verify_duration", 300))
    _verify_deadline = time.time() + max_duration
    print_fn(f"[*] Bat dau kiem tra vat ly (PIVOT) cho OOB: [bold]{alias}[/]")

    creds = get_all_credentials(cfg, oob_ip)
    own_hostname, working_cred = None, creds[0]
    for c in creds:
        try:
            tn = connect_auto(oob_ip, cfg.get("ssh_port", 22), cfg["telnet_port"], c["username"], c["password"], c["enable_password"], timeout=6)
            tn.write("\r\n")
            prmpt = tn.read_until(["#", ">", "cli->"], timeout=3)
            save_working_credential(oob_ip, c)
            if "cli->" in prmpt:
                m_h = re.search(r'Welcome to [^<]+<([^>]+)>', prmpt)
                own_hostname = m_h.group(1) if m_h else None
            else:
                own_hostname = fetch_hostname(tn)
            tn.close(); working_cred = c; break
        except Exception: pass
    own_hostname_clean = (own_hostname or "").strip().lower()

    results = [] 
    session = None

    def get_session(vendor):
        nonlocal session
        if session is None:
            session = connect_auto(oob_ip, cfg.get("ssh_port", 22), cfg["telnet_port"], working_cred["username"], working_cred["password"], working_cred["enable_password"], timeout=8)
            if vendor == "vertiv":
                session.write("cd access/")
                session.read_until("cli->", timeout=3)
            else:
                session.write("terminal length 0")
                session.read_until("#", timeout=2)
        return session
        
    def reset_session():
        nonlocal session
        if session: session.close()
        session = None

    def check_port_via_oob(t_ip, t_port, proto, vendor, t_desc):
        try: s = get_session(vendor)
        except Exception: raise RuntimeError("Khong the ket noi toi OOB")
        
        out = ""
        if vendor == "vertiv": 
            cmd = f"connect {t_desc}"
            s.write(cmd)
            
            # Buoc 1: Cho xem thiet bi hoi Pass hay vao thang/xuat hien Prompt/Hot key
            out_tmp = s.read_until(["assword:", "Password:", "Type the hot key", "cli->"], timeout=5)
            out += out_tmp
            
            if "assword:" in out_tmp or "Password:" in out_tmp:
                v_pass = cfg.get("vertiv_connect_password", "")
                s.write(v_pass)
                # Buoc 2: Cho xac thuc xong (co the ra Hot key, Prompt hoac quay ve cli->)
                out += s.read_until(["Type the hot key", "cli->", "login:", "Username:", "Password:"], timeout=12)
                
            # Đọc nốt dòng chứa Hot key để bỏ qua ký tự '>' trong <CTRL>Z
            if "Type the hot key" in out:
                out += s.read_until(["\n"], timeout=2)
                
            # Buoc 3: Session da mo -> Nghi 1s de on dinh, roi go Enter de trigger prompt neu vao thang
            time.sleep(1.0)
            s.write("") 
            
            # Xử lý trường hợp Vertiv báo Data Buffering Suspended cần Enter thêm
            out_tmp_buf = s.read_until(["Data Buffering Suspended"], timeout=1.5)
            out += out_tmp_buf
            if "Data Buffering Suspended" in out_tmp_buf:
                time.sleep(0.5)
                s.write("")
                
            time.sleep(0.5)
            s.write("") 
            
            # Đã bổ sung đầy đủ các trường hợp Prompt >, # cho thiết bị vào thẳng root
            out += s.read_until(["login:", "Username:", "Password:", ">", "#", "cli->", "%"], timeout=6)
            
        else: 
            cmd = f"ssh -l admin {t_ip}" if proto == "ssh" else f"telnet {t_ip} {t_port}"
            s.write(cmd)
            time.sleep(float(cfg.get("verify_wait_after_connect", 1.5))) 
            s.write("") 
            s.write("") 
            out += s.read_until([">", "#", "login:", "Username:", "Password:", "Connection refused", "refused", "unknown"], timeout=5)
        
        # Buoc 4: Thoat phien ket noi ve lai Vertiv CLI
        try:
            if vendor == "vertiv":
                s.write_raw(b"\x1a") # Gui Ctrl+Z de thoat phien console
                time.sleep(0.5)
                reset_session()
            else:
                s.write_raw(b"\x1ex")
                if "#" in s.read_until("#", timeout=2) or "#" in out:
                    s.write("disconnect")
                    if "[confirm]" in s.read_until(["[confirm]", "#", "No connection"], timeout=2): s.write(""); s.read_until("#", timeout=2)
                else: reset_session()
        except Exception: reset_session()
        
        # Buoc 5: Xoa "Bong ma Password" va chuoi noi dung Vertiv thua truoc do
        if vendor == "vertiv":
            idx = out.rfind("Type the hot key")
            if idx != -1:
                idx_nl = out.find("\n", idx)
                out = out[idx_nl:] if idx_nl != -1 else out[idx:]
                    
        return out

    def clear_line_via_oob(t_port, v_vendor):
        if v_vendor == "vertiv": return False
        try:
            s = get_session(v_vendor)
            line_num = t_port - 2000
            if line_num <= 0: return False
            s.write(f"clear line {line_num}")
            out = s.read_until(["[confirm]", "#"], timeout=3)
            if "[confirm]" in out:
                s.write("")
                s.read_until("#", timeout=2)
            return True
        except Exception:
            reset_session()
            return False

    def extract_hostname(output: str) -> str:
        output_clean = _ANSI_STRIP_RE.sub('', output)
        output_clean = re.sub(r'[\x00-\x09\x0b\x0c\x0e-\x1f\x7f]', '', output_clean)
        
        auth_seen = False
        lines = [l.strip() for l in output_clean.splitlines() if l.strip()]
        
        for line in reversed(lines):
            # Bo qua cac dong lenh hoac thong bao ket noi cua Vertiv/Network/OS
            if any(x in line for x in ["telnet ", "ssh ", "Trying ", "Open", "Connection refused", "disconnect", "clear line", "Type the hot key", "cli->", "Data Buffering Suspended", "{master"]): 
                continue
            
            # 1. Kiem tra Prompt truc tiep (VD: HCM-ROUTER-01> hoac HCM-ROUTER-01#)
            m_prompt = _HOSTNAME_PROMPT_RE.search(line)
            if m_prompt: 
                h = m_prompt.group(1)
                # Loai tru prompt noi bo cua Vertiv
                if h.lower() not in ["cli", "cli-", "access", "admin", "root"]: 
                    return h

            # 2. Kiem tra dang nhap kieu Unix/Linux (hostname login:)
            m_login = _HOSTNAME_LOGIN_RE.search(line)
            if m_login: return m_login.group(1)
            
            # 3. Kiem tra đi kèm he dieu hanh FreeBSD / Linux
            m_bsd = _HOSTNAME_BSD_RE.search(line)
            if m_bsd: return m_bsd.group(1)
                
            m_bsd_new = re.search(r'(?:FreeBSD|Linux|NetBSD|OpenBSD).*?\(([A-Za-z0-9_\-\.]+)\)', line, re.IGNORECASE)
            if m_bsd_new: return m_bsd_new.group(1)
                    
            if re.search(r'Username:|Password:|login:', line, re.IGNORECASE): 
                auth_seen = True
                
        return "AUTH_REQUIRED" if auth_seen else None

    for key, opt in options.items():
        desc = opt.get("description", "")
        if not desc: continue
        target_ip, port, proto, vendor = opt.get("ip"), opt.get("port", 23), opt.get("protocol", "telnet"), opt.get("vendor", "cisco")
        act_host, note_parts = None, []

        try: act_host = extract_hostname(check_port_via_oob(target_ip, port, proto, vendor, desc))
        except Exception: pass
            
        if not act_host:
            if port > 2000 and vendor != "vertiv":
                print_fn(f"[yellow][!][/] {alias} (Opt {key}): Dang clear line {port - 2000}...")
                note_parts.append(f"Da thu clear line {port - 2000}")
                if clear_line_via_oob(port, vendor):
                    time.sleep(2) 
                    try: act_host = extract_hostname(check_port_via_oob(target_ip, port, proto, vendor, desc))
                    except Exception: pass
                else: note_parts.append("Khong clear duoc line")
            else: note_parts.append("Port la Direct/Vertiv, bo qua clear line")

        note = "; ".join(note_parts)
        if not act_host:
            print_fn(f"[dim][-][/] {alias} (Opt {key}): TIMEOUT hoac Loi mang")
            results.append({"key": key, "status": "TIMEOUT", "act_host": None, "desc": desc, "port": port, "note": note})
        elif act_host == "AUTH_REQUIRED":
            print_fn(f"[yellow][?][/] {alias} (Opt {key}): Thiet bi yeu cau dang nhap")
            results.append({"key": key, "status": "YEU CAU DANG NHAP", "act_host": None, "desc": desc, "port": port, "note": note})
        else:
            desc_clean, act_host_clean = _DESC_PREFIX_RE.sub('', desc).strip().lower(), act_host.strip().lower()
            if own_hostname_clean and act_host_clean == own_hostname_clean:
                print_fn(f"[dim][-][/] {alias} (Opt {key}): Khong pivot duoc (van o console OOB)")
                results.append({"key": key, "status": "KHONG PIVOT", "act_host": act_host, "desc": desc, "port": port, "note": note})
            elif act_host_clean == desc_clean or hostname_matches_description(act_host, desc_clean):
                print_fn(f"[green](OK)[/] {alias} (Opt {key}): Khop ({act_host})")
                results.append({"key": key, "status": "OK", "act_host": act_host, "desc": desc, "port": port, "note": note})
            else:
                msg_alarm = f"ALARM! Thiet bi noi line ([yellow]{act_host}[/]) khac description ([yellow]{desc_clean}[/]) tai Opt {key} tren {alias}!"
                print_fn(f"\a[bold red blink][ALARM] {msg_alarm}[/bold red blink]")
                sys.stdout.write('\a'); sys.stdout.flush()
                
                os.makedirs("alarms", exist_ok=True)
                try:
                    with file_lock:
                        with open("alarms/verify_alarms.log", "a", encoding="utf-8") as f:
                            f.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ALARM: {alias} (Opt {key}) - Thuc te '{act_host}' != Desc '{desc_clean}'\n")
                except Exception: pass
                
                results.append({"key": key, "status": "CANH BAO", "act_host": act_host, "desc": desc, "port": port, "note": note})

    if session: session.close()
    
    os.makedirs("verify-logs", exist_ok=True)
    ts_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = os.path.join("verify-logs", f"Verify_{alias}_{ts_str}.log")
    with file_lock:
        with open(log_path, "w", encoding="utf-8") as f: f.write(_build_verify_report(alias, oob_ip, own_hostname, results))
        try:
            with open(log_path.replace(".log", ".json"), "w", encoding="utf-8") as f: json.dump(results, f, ensure_ascii=False, indent=2)
        except Exception: pass
    return results

def _thread_verify_only(cfg, alias, ip, snapshot, pfx="", prog_state=None):
    if prog_state:
        with prog_state["lock"]:
            prog_state["started"] += 1
            idx = prog_state["started"]
            tot = prog_state["total"]
            pct = (idx / tot) * 100 if tot else 0
            pfx = f"[{idx}/{tot} {pct:.0f}%]"
            
    log_verify(f"{pfx} [*] Bat dau kiem tra vat ly (PIVOT) cho OOB: [bold]{alias}[/]")
    def _print_fn(msg): log_verify(f"{pfx} {msg}")
    run_deep_verify(cfg, alias, ip, snapshot, print_fn=_print_fn)
    
    if prog_state:
        with prog_state["lock"]:
            prog_state["completed"] += 1
            c_idx = prog_state["completed"]
            tot = prog_state["total"]
            c_pct = (c_idx / tot) * 100 if tot else 0
            pfx_done = f"[{c_idx}/{tot} {c_pct:.0f}%]"
    else:
        pfx_done = pfx
    log_verify(f"{pfx_done} [green](OK)[/] Hoan thanh Verify cho OOB: [bold]{alias}[/]\n")

# ---------------------------------------------------------------------------
# Manual Push
# ---------------------------------------------------------------------------
def process_push_and_reverify(cfg, alias, oob_ip, baseline, verify_results, print_fn=None):
    if print_fn is None: print_fn = log_verify
    warnings = [r for r in verify_results if r["status"] == "CANH BAO"]
    if not warnings: return

    updates_list, push_log_entries = [], []
    mn, dn, _ = get_options_by_host(cfg["baseline_db"], "baseline_menu", oob_ip)
    if not mn: return

    for w in warnings:
        key, act_host, opt = w["key"], w["act_host"], baseline.get(w["key"], {})
        target_ip = opt.get("ip")
        
        with db_lock:
            conn = _init_db(cfg["baseline_db"], "baseline_menu")
            cur = conn.execute("SELECT host FROM baseline_menu WHERE target_ip=?", (target_ip,))
            hosts = {row[0] for row in cur.fetchall()}
            conn.close()
        
        if len(hosts - {oob_ip}) > 0:
            print_fn(f"[yellow][!][/] {alias} (Opt {key}): Nghi ngo trung IP tren nhieu OOB. Bo qua tu dong sua.")
            continue
            
        new_desc = f"----> {act_host}"
        real_menu_name, real_key = opt.get("_menu_name"), opt.get("_raw_key")
        if not real_menu_name or not real_key: continue
        updates_list.append((real_menu_name, real_key, new_desc))
        push_log_entries.append({"key": key, "real_menu_name": real_menu_name, "real_key": real_key, "target_ip": target_ip, "old": w["desc"], "new": new_desc})

    if not updates_list: return
    print_fn(f"[*] Dang PUSH thuc thi sua loi {len(updates_list)} option cho {alias}...")
    
    success = False
    for c in get_all_credentials(cfg, oob_ip):
        success = push_menu_descriptions(oob_ip, cfg.get("ssh_port", 22), cfg["telnet_port"], c["username"], c["password"], c["enable_password"], updates_list, timeout=10)
        if success:
            save_working_credential(oob_ip, c)
            break

    if not success:
        print_fn(f"[red][LOI][/] {alias}: Push cau hinh that bai (Cisco tu choi hoac sai authen).")
        return

    os.makedirs("push-logs", exist_ok=True)
    ts_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    with file_lock:
        with open(os.path.join("push-logs", f"Push_{alias}_{ts_str}.log"), "w", encoding="utf-8") as f:
            f.write(f"=== PUSH LOG THU CONG: {alias} ({oob_ip}) ===\nThoi gian: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            for entry in push_log_entries:
                baseline[entry["key"]]["description"] = entry["new"] 
                f.write(f"- Option [{entry['key']}] (Target IP: {entry['target_ip']}):\n  + Cu : {entry['old']}\n  + Moi: {entry['new']}\n")

    save_options(cfg["baseline_db"], "baseline_menu", oob_ip, mn, dn, baseline)
    print_fn(f"[green](OK)[/] Da cap nhat cau hinh vao Switch & cap nhat lai Baseline DB.")
    print_fn(f"[*] Tu dong Re-Verify lai cac option vua sua tren {alias}...")
    run_deep_verify(cfg, alias, oob_ip, {entry["key"]: baseline[entry["key"]] for entry in push_log_entries}, print_fn=print_fn)

def manual_push_devices(cfg):
    targets_input = _con.input("  [cyan]Nhap IP/Alias can kiem tra de PUSH (cach nhau dau phay, de trong quet TAT CA)[/]: ").strip()
    all_hosts = load_ip_list(cfg["ip_list"])
    if not all_hosts: return _con.print("  [yellow][!][/] Danh sach IP dang trong.")
    
    hosts_to_scan = []
    if not targets_input: hosts_to_scan = all_hosts
    else:
        target_list = [t.strip().lower() for t in targets_input.split(",")]
        for ip, alias in all_hosts:
            if ip.lower() in target_list or alias.lower() in target_list: hosts_to_scan.append((ip, alias))
            
    if not hosts_to_scan: return
    _con.print(f"\n  [green][*] Bat dau Verify & Push thu cong cho {len(hosts_to_scan)} thiet bi...[/]")
    def cli_print(msg): _con.print(f"    {msg}")
    
    for idx, (ip, alias) in enumerate(hosts_to_scan, 1):
        pct = (idx / len(hosts_to_scan)) * 100
        pfx = f"[{idx}/{len(hosts_to_scan)} {pct:.0f}%]"

        _con.print(f"\n  [cyan]{pfx} [MANUAL PUSH][/] [bold]{alias}[/] ({ip}) ...")
        _mn, _dn, baseline = get_options_by_host(cfg["baseline_db"], "baseline_menu", ip)
        if not baseline:
            _con.print(f"  [yellow][!][/] OOB nay chua co Baseline. Vui long quet cau hinh (Option 7) truoc!")
            continue
            
        vendor = next(iter(baseline.values())).get("vendor", "cisco") if baseline else "cisco"
        if vendor == "vertiv":
            _con.print(f"  [yellow][!][/] Thiet bi {alias} la Vertiv. Tinh nang Push Config chua duoc ho tro cho hang nay!")
            run_deep_verify(cfg, alias, ip, baseline, cli_print)
            continue
            
        results = run_deep_verify(cfg, alias, ip, baseline, cli_print)
        if any(r["status"] == "CANH BAO" for r in results):
            ans = _con.input(f"\n  [bold yellow]Phat hien sai lech tren {alias}. Ban co chac chan PUSH de sua Description? (y/N)[/]: ").strip().lower()
            if ans == 'y': process_push_and_reverify(cfg, alias, ip, baseline, results, print_fn=cli_print)
            else: _con.print(f"  [dim]Da huy PUSH cho {alias}.[/]")
        else:
            _con.print(f"  [green](OK)[/] Khong co sai lech nao can sua cho {alias}.")

# ---------------------------------------------------------------------------
# Daemon Thread
# ---------------------------------------------------------------------------
def run_verify_daemon(config_path):
    cfg = load_config(config_path)
    log_verify(f"[green][START][/] Khoi dong Verify vat ly - lich: {_describe_verify_schedule(cfg)}.")
    time.sleep(15)
    while True:
        cfg = load_config(config_path)
        
        if cfg.get("auto_verify", True):
            hosts = load_ip_list(cfg["ip_list"])
            if hosts:
                valid_hosts = []
                for ip, alias in hosts:
                    _mn, _dn, baseline = get_options_by_host(cfg["baseline_db"], "baseline_menu", ip)
                    if baseline: valid_hosts.append((ip, alias, baseline))
                    
                if valid_hosts:
                    prog_state = {"lock": threading.Lock(), "started": 0, "completed": 0, "total": len(valid_hosts)}
                    with action_lock: 
                        with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
                            for ip, alias, baseline in valid_hosts:
                                executor.submit(_thread_verify_only, cfg, alias, ip, baseline, "", prog_state)
        else:
            log_verify("[dim][zzz] Tinh nang Verify ngam dang bi TAT trong cau hinh. Dang cho...[/]")

        schedule_mode = cfg.get("verify_schedule_mode", "interval")
        if schedule_mode in ("daily", "weekly"):
            next_run = compute_next_scheduled_run(schedule_mode, cfg.get("verify_schedule_time", "01:00"), cfg.get("verify_schedule_weekday", "mon"))
            sleep_seconds = max(1, int((next_run - datetime.now()).total_seconds()))
            log_verify(f"[dim][zzz] Lan Verify tiep theo: {next_run.strftime('%Y-%m-%d %H:%M')} (con {sleep_seconds}s)...[/]")
        else:
            sleep_seconds = cfg.get("verify_interval", 3600)
            log_verify(f"[dim][zzz] Dang cho {sleep_seconds}s cho dot Verify tiep theo...[/]")
        time.sleep(sleep_seconds)

def _scan_wait(cfg):
    mode = cfg.get("scan_schedule_mode", "interval")
    if mode in ("daily", "weekly"):
        next_run = compute_next_scheduled_run(mode, cfg.get("scan_schedule_time", "01:00"), cfg.get("scan_schedule_weekday", "mon"))
        sleep_seconds = max(1, int((next_run - datetime.now()).total_seconds()))
        log_oob(f"[dim][zzz] Lan Thu thap tiep theo: {next_run.strftime('%Y-%m-%d %H:%M')} (con {sleep_seconds}s)...[/]")
    else:
        sleep_seconds = cfg["interval"]
        log_oob(f"[dim][zzz] Dang cho {sleep_seconds}s de quet lai...[/]")
    time.sleep(sleep_seconds)

def run_daemon(cfg, config_path=None):
    global _live_ui
    _con.print(Panel("[bold green]OOB MONITOR DAEMON[/]\n[dim]Dang giam sat lien tuc. Nhan Ctrl+C de dung.[/]", border_style="green"))
    
    creds = get_all_credentials(cfg)
    if not any(c.get("password") for c in creds):
        _con.print("[red][!][/] Chua cau hinh password! Vui long cau hinh truoc.")
        return

    update_ui()
    threading.Thread(target=run_verify_daemon, args=(config_path or CONFIG_FILE_DEFAULT,), daemon=True).start()
    threading.Thread(target=_daemon_heartbeat_loop, daemon=True).start()
    
    with Live(layout, refresh_per_second=4, screen=False) as live:
        _live_ui = live
        log_oob(f"[green][START][/] Khoi dong Thu thap cau hinh - lich: {_describe_scan_schedule(cfg)}.")
        
        try:
            while True:
                hosts = load_ip_list_cached(cfg["ip_list"])
                if not hosts:
                    log_oob("[yellow][!][/] Danh sach IP trong. Doi them thiet bi...")
                    _scan_wait(cfg)
                    continue

                total_hosts = len(hosts)
                started_hosts = 0
                scan_prog_lock = threading.Lock()
                pending_verifies = []

                def _scan_single_daemon(ip, alias):
                    nonlocal started_hosts
                    with scan_prog_lock:
                        started_hosts += 1
                        idx = started_hosts
                        pct = (idx / total_hosts) * 100
                    pfx = f"[{idx}/{total_hosts} {pct:.0f}%]"

                    log_oob(f"[cyan]{pfx} [PING][/] [bold]{alias}[/] ({ip}) ...")
                    alive = ping_host(ip)
                    save_device_status(ip, alias=alias, ping=alive)
                    if not alive:
                        log_oob(f"[red]{pfx} [!][/] {alias} ({ip}): Khong ping duoc. Bo qua vong nay.")
                        return

                    log_oob(f"[cyan]{pfx} [SCAN][/] [bold]{alias}[/] ({ip}) ...")
                    try: hostname, menu_name, snapshot, menu_state = poll_host_multi(ip, cfg, timeout=cfg.get("interval", 30))
                    except Exception as exc:
                        save_device_status(ip, alias=alias, menu_state="conn_failed")
                        log_oob(f"[red]{pfx} [LOI][/] {alias} ({ip}): {exc}")
                        return

                    save_device_status(ip, alias=alias, menu_state=menu_state)
                    if menu_state == "fetch_failed":
                        log_oob(f"[yellow]{pfx} [!][/] {alias}: Da ket noi nhung KHONG lay duoc thong tin menu.")
                        return
                    if menu_state == "no_menu" or not menu_name or not snapshot:
                        log_oob(f"[yellow]{pfx} [!][/] {alias}: Thiet bi khong co cau hinh menu.")
                        return

                    save_options(cfg["snapshot_db"], "snapshot_menu", ip, menu_name, hostname, snapshot)
                    _mn, _dn, baseline = get_options_by_host(cfg["baseline_db"], "baseline_menu", ip)

                    if baseline is None:
                        with ui_lock:
                            live.stop()
                            _con.print(f"\n  [yellow]{pfx} [?][/] Chua co baseline cho [bold]{alias}[/] ({ip}).")
                            print_options(snapshot)
                            save_options(cfg["baseline_db"], "baseline_menu", ip, menu_name, hostname, snapshot)
                            log_baseline_change(alias, ip, "TAO MOI BASELINE")
                            _con.print(f"  [green]{pfx} (OK)[/] Da TU DONG luu baseline cho {alias}.")
                            live.start()
                        pending_verifies.append((alias, ip, snapshot, pfx))
                        return

                    if options_equal(baseline, snapshot):
                        log_oob(f"[green]{pfx} (OK)[/] {alias}: Khop voi baseline ({len(snapshot)} option).")
                        return

                    with ui_lock:
                        live.stop()
                        _con.rule(f"[bold red]CANH BAO  {alias} ({ip}) KHAC baseline![/]", style="red")
                        print_diff(baseline, snapshot)
                        save_options(cfg["baseline_db"], "baseline_menu", ip, menu_name, hostname, snapshot)
                        log_baseline_change(alias, ip, "CAP NHAT BASELINE")
                        _con.print(f"  [green]{pfx} (OK)[/] Da TU DONG cap nhat baseline moi cho {alias}.")
                        live.start()
                    pending_verifies.append((alias, ip, snapshot, pfx))

                with action_lock: 
                    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
                        futures = [executor.submit(_scan_single_daemon, ip, alias) for ip, alias in hosts]
                        concurrent.futures.wait(futures)

                    if cfg.get("auto_verify", True) and pending_verifies:
                        prog_state = {"lock": threading.Lock(), "started": 0, "completed": 0, "total": len(pending_verifies)}
                        with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
                            for alias, ip, snap, _ in pending_verifies:
                                executor.submit(_thread_verify_only, cfg, alias, ip, snap, "", prog_state)

                _scan_wait(cfg)
        except KeyboardInterrupt: pass
    _con.print("\n[yellow][STOP][/] Da nhan Ctrl+C. Dung Daemon.")

def scan_specific_devices(cfg):
    targets_input = _con.input("  [cyan]Nhap IP/Alias can quet (cach nhau dau phay, de trong quet TAT CA)[/]: ").strip()
    all_hosts = load_ip_list(cfg["ip_list"])
    if not all_hosts: return _con.print("  [yellow][!][/] Danh sach IP hien dang trong.")
    
    hosts_to_scan = []
    if not targets_input: hosts_to_scan = all_hosts
    else:
        target_list = [t.strip().lower() for t in targets_input.split(",")]
        for ip, alias in all_hosts:
            if ip.lower() in target_list or alias.lower() in target_list: hosts_to_scan.append((ip, alias))
    if not hosts_to_scan: return
    
    _con.print(f"\n  [green][*] Bat dau quet C?u h?nh ?A LU?NG {len(hosts_to_scan)} thiet bi...[/]")
    
    total_hosts = len(hosts_to_scan)
    started_hosts = 0
    scan_prog_lock = threading.Lock()
    pending_verifies = []

    def _scan_single_cli(ip, alias):
        nonlocal started_hosts
        with scan_prog_lock:
            started_hosts += 1
            idx = started_hosts
            pct = (idx / total_hosts) * 100
        pfx = f"[{idx}/{total_hosts} {pct:.0f}%]"

        with ui_print_lock: _con.print(f"\n  [cyan]{pfx} [PING][/] [bold]{alias}[/] ({ip}) ...")
        alive = ping_host(ip)
        save_device_status(ip, alias=alias, ping=alive)
        if not alive:
            with ui_print_lock: _con.print(f"  [red]{pfx} [!][/] {alias} ({ip}): Khong ping duoc. Bo qua.")
            return

        with ui_print_lock: _con.print(f"  [cyan]{pfx} [SCAN CONFIG][/] [bold]{alias}[/] ({ip}) ...")
        try: hostname, menu_name, snapshot, menu_state = poll_host_multi(ip, cfg, timeout=10)
        except Exception as exc:
            save_device_status(ip, alias=alias, menu_state="conn_failed")
            with ui_print_lock: _con.print(f"  [red]{pfx} [LOI][/] {alias} ({ip}): {exc}")
            return

        save_device_status(ip, alias=alias, menu_state=menu_state)
        if menu_state == "fetch_failed":
            with ui_print_lock: _con.print(f"  [yellow]{pfx} [!][/] {alias}: Da ket noi nhung KHONG lay duoc thong tin menu.")
            return
        if menu_state == "no_menu" or not menu_name or not snapshot:
            with ui_print_lock: _con.print(f"  [yellow]{pfx} [!][/] {alias}: Thiet bi khong co cau hinh menu.")
            return

        save_options(cfg["snapshot_db"], "snapshot_menu", ip, menu_name, hostname, snapshot)
        _mn, _dn, baseline = get_options_by_host(cfg["baseline_db"], "baseline_menu", ip)

        if baseline is None:
            with ui_print_lock:
                _con.print(f"  [yellow]{pfx} [?][/] Chua co baseline cho [bold]{alias}[/] ({ip}).")
                save_options(cfg["baseline_db"], "baseline_menu", ip, menu_name, hostname, snapshot)
                log_baseline_change(alias, ip, "TAO MOI BASELINE")
                _con.print(f"  [green]{pfx} (OK)[/] Da TU DONG luu baseline cho {alias}.")
            pending_verifies.append((alias, ip, snapshot, pfx))
            return

        if options_equal(baseline, snapshot):
            with ui_print_lock: _con.print(f"  [green]{pfx} (OK)[/] {alias}: Khop voi baseline ({len(snapshot)} option).")
            return

        with ui_print_lock:
            _con.rule(f"[bold red]CANH BAO  {alias} ({ip}) KHAC baseline![/]", style="red")
            print_diff(baseline, snapshot)
            save_options(cfg["baseline_db"], "baseline_menu", ip, menu_name, hostname, snapshot)
            log_baseline_change(alias, ip, "CAP NHAT BASELINE")
            _con.print(f"  [green]{pfx} (OK)[/] Da TU DONG cap nhat baseline moi cho {alias}.")
        pending_verifies.append((alias, ip, snapshot, pfx))

    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
        futures = [executor.submit(_scan_single_cli, ip, alias) for ip, alias in hosts_to_scan]
        concurrent.futures.wait(futures)

    if cfg.get("auto_verify", True) and pending_verifies:
        prog_state = {"lock": threading.Lock(), "started": 0, "completed": 0, "total": len(pending_verifies)}
        with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
            for alias, ip, snap, _ in pending_verifies:
                executor.submit(_thread_verify_only, cfg, alias, ip, snap, "", prog_state)

def verify_specific_devices(cfg):
    targets_input = _con.input("  [cyan]Nhap IP/Alias can Verify (cach nhau dau phay, de trong quet TAT CA)[/]: ").strip()
    all_hosts = load_ip_list(cfg["ip_list"])
    if not all_hosts: return _con.print("  [yellow][!][/] Danh sach IP hien dang trong.")
    
    hosts_to_scan = []
    if not targets_input: hosts_to_scan = all_hosts
    else:
        target_list = [t.strip().lower() for t in targets_input.split(",")]
        for ip, alias in all_hosts:
            if ip.lower() in target_list or alias.lower() in target_list: hosts_to_scan.append((ip, alias))

    if not hosts_to_scan: return
    _con.print(f"\n  [green][*] Bat dau Verify vat ly tuc thi {len(hosts_to_scan)} thiet bi...[/]")
    
    valid_hosts = []
    for ip, alias in hosts_to_scan:
        _mn, _dn, baseline = get_options_by_host(cfg["baseline_db"], "baseline_menu", ip)
        if not baseline:
            _con.print(f"  [yellow][!][/] OOB nay chua co Baseline. Vui long quet cau hinh truoc (Option 7)!")
        else:
            valid_hosts.append((ip, alias, baseline))
            
    if not valid_hosts: return
    
    prog_state = {"lock": threading.Lock(), "started": 0, "completed": 0, "total": len(valid_hosts)}

    def _worker(alias, ip, baseline):
        with prog_state["lock"]:
            prog_state["started"] += 1
            idx = prog_state["started"]
            tot = prog_state["total"]
            pct = (idx / tot) * 100 if tot else 0
            pfx = f"[{idx}/{tot} {pct:.0f}%]"
            
        with ui_print_lock:
            _con.print(f"\n  [cyan]{pfx} [VERIFY][/] Dang Verify: [bold]{alias}[/] ({ip}) ...")
            
        def cli_print(msg):
            with ui_print_lock: _con.print(f"    {pfx} {msg}")
            
        run_deep_verify(cfg, alias, ip, baseline, cli_print)
        
        with prog_state["lock"]:
            prog_state["completed"] += 1
            c_idx = prog_state["completed"]
            tot = prog_state["total"]
            c_pct = (c_idx / tot) * 100 if tot else 0
            pfx_done = f"[{c_idx}/{tot} {c_pct:.0f}%]"
        with ui_print_lock:
            _con.print(f"  [green]{pfx_done} [OK][/] Da hoan thanh Verify: [bold]{alias}[/]")
            
    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
        for ip, alias, baseline in valid_hosts:
            executor.submit(_worker, alias, ip, baseline)

# ---------------------------------------------------------------------------
# Import / Export / TIM KIEM
# ---------------------------------------------------------------------------
def _parse_verify_logs_for_status(max_age_hours: float = 24.0) -> dict:
    log_dir = "verify-logs"
    if not os.path.exists(log_dir): return {}
    cutoff = time.time() - max_age_hours * 3600
    alias_files: dict = {}
    for fname in os.listdir(log_dir):
        if not fname.endswith('.json') or not fname.startswith('Verify_'): continue
        body = fname[len('Verify_'):-len('.json')]
        if len(body) < 17: continue
        alias, fpath = body[:-16], os.path.join(log_dir, fname)
        mtime = os.path.getmtime(fpath)
        if mtime < cutoff: continue
        if alias not in alias_files or mtime > alias_files[alias][1]: alias_files[alias] = (fpath, mtime)

    STATUS_MAP = {"OK": "OK", "CANH BAO": "CANH BAO", "KO PIVOT": "KHONG PIVOT", "TIMEOUT": "TIMEOUT", "YC DANG NHAP": "YEU CAU DANG NHAP"}
    result: dict = {}
    for alias, (fpath, _) in alias_files.items():
        try:
            with open(fpath, "r", encoding="utf-8") as f: data = json.load(f)
        except Exception: continue
        for item in data:
            opt_key = item.get("key")
            if not opt_key: continue
            status_raw = item.get("status", "")
            act_host = item.get("act_host")
            result[(alias, opt_key)] = {"status": STATUS_MAP.get(status_raw, status_raw), "act_host": act_host if act_host not in ('-', '') else None}
    return result

def search_device(cfg):
    query = input("  Nhap IP/Ten thiet bi: ").strip().lower()
    if not query: return
    hosts = load_ip_list(cfg["ip_list"])
    found = []
    
    verify_st = _parse_verify_logs_for_status(max_age_hours=24.0 * 30)
    
    for ip, alias in hosts:
        _mn, dn, source = get_options_by_host(cfg["baseline_db"], "baseline_menu", ip)
        if source is None: _mn, dn, source = get_options_by_host(cfg["snapshot_db"], "snapshot_menu", ip)
        if not source: continue
        dn = dn or alias
        
        for key, entry in source.items():
            act_host = verify_st.get((alias, key), {}).get("act_host", "") or ""
            score = 0
            if query == act_host.lower(): score = 100
            elif query in act_host.lower(): score = 90
            elif query == dn.lower(): score = 80
            elif query in dn.lower(): score = 70
            elif query in entry.get("description", "").lower(): score = 60
            elif query in entry.get("ip", "").lower(): score = 50
            elif query == key.lower(): score = 40
            
            if score > 0:
                found.append((score, ip, alias, dn, key, entry, act_host))
                
    if not found: return _con.print(f"\n  [yellow]Khong tim thay thiet bi nao![/]")
    found.sort(key=lambda x: x[0], reverse=True)
    
    _con.print(f"\n  [green][*] Tim thay {len(found)} ket qua (Uu tien: Hostname thuc > Hostname OOB > Description):[/]")
    for score, ip, alias, dn, key, entry, act_host in found:
        pr = _norm_proto(entry.get("protocol"))
        pt = entry.get("port", 22 if pr == "ssh" else 23)
        act_str = f" [bold green](Thuc te: {act_host})[/]" if act_host else ""
        _con.print(f"    [cyan]->[/] OOB: [bold]{alias}[/] ({ip} - host: {dn}) | Opt [bold cyan]{key}[/] {entry.get('description', '')}{act_str} [dim]-> {pr}://{entry['ip']}:{pt}[/]")

def view_latest_verify_log():
    log_dir = "verify-logs"
    if not os.path.exists(log_dir): return _con.print("\n  [yellow][!][/] Chua co thu muc 'verify-logs'.")
    files = [os.path.join(log_dir, f) for f in os.listdir(log_dir) if f.endswith('.log')]
    if not files: return _con.print("\n  [yellow][!][/] Thu muc 'verify-logs' dang trong.")
    latest_file = max(files, key=os.path.getmtime)
    try:
        with open(latest_file, "r", encoding="utf-8") as f: content = f.read()
        _con.print(Panel(content, title=f"[bold magenta] [ LOG VERIFY GAN NHAT: {os.path.basename(latest_file)} ] [/]", border_style="magenta", padding=(1, 2)))
    except Exception as e: _con.print(f"  [red][LOI][/] {e}")

def list_devices(cfg):
    hosts = load_ip_list(cfg["ip_list"])
    if not hosts: return _con.print(f"\n  [yellow][!][/] Danh sach trong.")
    table = Table(title="[bold]DANH SACH THIET BI OOB[/]", box=rbox.ASCII, border_style="cyan", header_style="bold cyan")
    table.add_column("Alias", style="bold cyan", min_width=12)
    table.add_column("IP", min_width=16)
    table.add_column("Hostname", min_width=14)
    table.add_column("Baseline", justify="center", min_width=9)
    table.add_column("Cap nhat luc", min_width=20)
    checked_count = unchecked_count = 0
    for ip, alias in hosts:
        _mn, device_name, baseline = get_options_by_host(cfg["baseline_db"], "baseline_menu", ip)
        updated_at = get_updated_at_by_host(cfg["baseline_db"], "baseline_menu", ip)
        if device_name is None: _mn, device_name, _ = get_options_by_host(cfg["snapshot_db"], "snapshot_menu", ip)
        if baseline: bl = "[green](OK) Co[/]"; checked_count += 1
        else: bl = "[yellow](X) Chua[/]"; unchecked_count += 1
        table.add_row(alias, ip, device_name or "[dim](chua ro)[/]", bl, updated_at or "[dim]-[/]")
    _con.print(f"  [cyan]TONG KET: Co [bold]{len(hosts)}[/] thiet bi | Da check: [green bold]{checked_count}[/] | Chua check: [yellow bold]{unchecked_count}[/][/]")
    _con.print(table)

def view_baseline(cfg):
    hosts = load_ip_list(cfg["ip_list"])
    _con.print("\n" + "[bold cyan]BASELINE (CHUAN) DA LUU[/]".center(50, " "))
    found = False
    for ip, alias in hosts:
        mn, dn, baseline = get_options_by_host(cfg["baseline_db"], "baseline_menu", ip)
        if baseline is None: continue
        found = True
        upd = get_updated_at_by_host(cfg["baseline_db"], "baseline_menu", ip)
        _con.print(f"\n  [bold cyan]{alias}[/] ({ip})  host: [bold]{dn or alias}[/]  upd: [dim]{upd}[/]")
        print_options(baseline)
    if not found: _con.print("  [dim](Chua co baseline nao duoc xac nhan)[/]")

def _export_excel_template():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError: return _con.print("  [red][!][/] Thieu openpyxl. Chay: pip install openpyxl")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OOB_Import"
    for col, header in enumerate(["IP", "Alias (ten goi)"], 1):
        c = ws.cell(1, col, header)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill(fill_type="solid", fgColor="1F4E79")
        c.alignment = Alignment(horizontal="center", vertical="center")
    for r, (ip, alias) in enumerate([("192.168.1.1", "OOB-HCM-01"), ("192.168.1.2", "OOB-HCM-02")], 2):
        ws.cell(r, 1, ip); ws.cell(r, 2, alias)
    ws.column_dimensions['A'].width = 18; ws.column_dimensions['B'].width = 24
    tpl = "oob_import_template.xlsx"
    wb.save(tpl)
    _con.print(f"  [green](OK)[/] Da tao file mau: [bold]{tpl}[/]")
    _con.print("      Dien IP vao cot A, ten alias vao cot B.")

def import_from_excel(cfg):
    try: import openpyxl
    except ImportError: return _con.print("  [red][!][/] Thieu thu vien openpyxl. (pip install openpyxl)")
    want_tpl = _con.input("  [cyan]Xuat file mau Excel de tham khao dinh dang? (y/N)[/]: ").strip().lower()
    if want_tpl == 'y': return _export_excel_template()
    file_path = _con.input("  [cyan]Duong dan file Excel (.xlsx)[/]: ").strip()
    if not file_path or not os.path.exists(file_path): return _con.print(f"  [red][!][/] Khong tim thay file.")
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e: return _con.print(f"  [red][!][/] Khong doc duoc: {e}")
    existing_ips = {h[0] for h in load_ip_list(cfg["ip_list"])}
    added = skipped_dup = skipped_invalid = 0
    try:
        with open(cfg["ip_list"], "a", encoding="utf-8") as f:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None: continue
                ip_raw = str(row[0]).strip()
                alias = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ip_raw
                if not _IP_RE.match(ip_raw): skipped_invalid += 1; continue
                if ip_raw in existing_ips: skipped_dup += 1; continue
                f.write(f"{ip_raw} {alias}\n")
                existing_ips.add(ip_raw)
                added += 1
    finally:
        try: wb.close()
        except: pass
    _con.print(f"\n  [green](OK)[/] Da them {added}, bo qua {skipped_dup} trung, bo qua {skipped_invalid} khong hop le.")

def export_menu_report(cfg):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError: return _con.print("  [red][!][/] Thieu thu vien openpyxl. (pip install openpyxl)")

    hosts = load_ip_list(cfg["ip_list"])
    if not hosts: return _con.print("  [yellow][!][/] Danh sach IP dang trong.")
    _con.print("  [cyan][*][/] Dang xuat bao cao Excel...")
    verify_st = _parse_verify_logs_for_status()
    dev_status = load_device_status()

    def _ping_label(ip):
        p = dev_status.get(ip, {}).get("ping")
        if p is True: return "Song (OK)"
        if p is False: return "Down (khong ping duoc)"
        return "Chua kiem tra"

    def _menu_state_label(ip):
        ms = dev_status.get(ip, {}).get("menu_state")
        if ms == "conn_failed": return "Loi ket noi/xac thuc"
        return MENU_STATE_LABELS.get(ms, "Chua kiem tra")

    C_MATCH, C_WRONG  = "C6EFCE", "FFC7CE"
    C_UNVER, C_NO_CON = "FFEB9C", "FFCC99"
    C_NO_DS           = "D9D9D9"
    C_HDR, C_HDR2     = "1F4E79", "2E75B6"

    def mk_fill(c): return PatternFill(fill_type="solid", fgColor=c)
    def mk_bdr():
        s = Side(style='thin', color='BFBFBF')
        return Border(left=s, right=s, top=s, bottom=s)

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Chi tiet"
    h1 = ["OOB IP", "OOB Alias", "Ping", "Trang thai Menu", "OOB Hostname", "Menu Name", "Option Key", "Description", "Target IP", "Target Port", "Protocol", "Vendor", "Desc Status", "Ghi chu"]
    for ci, h in enumerate(h1, 1):
        c = ws1.cell(1, ci, h); c.font = Font(bold=True, color="FFFFFF"); c.fill = mk_fill(C_HDR); c.alignment = Alignment(horizontal="center"); c.border = mk_bdr()
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(h1))}1"
    
    ri, summary = 2, []
    for ip, alias in hosts:
        ping_lbl, menu_state_lbl = _ping_label(ip), _menu_state_label(ip)
        mn, device_name, baseline = get_options_by_host(cfg["baseline_db"], "baseline_menu", ip)
        if baseline is None:
            for ci, val in enumerate([ip, alias, ping_lbl, menu_state_lbl, "(chua co baseline)", "", "", "", "", "", "", "", ""], 1): ws1.cell(ri, ci, val)
            ri += 1
            summary.append({"alias": alias, "ip": ip, "ping": ping_lbl, "menu_state": menu_state_lbl, "hn": "", "mn": "", "total": 0, "match": 0, "wrong": 0, "unverif": 0, "no_conn": 0, "no_desc": 0})
            continue

        hn = device_name or ""
        cnt = dict(match=0, wrong=0, unverif=0, no_conn=0, no_desc=0)
        start_ri = ri
        for opt_key in sorted(baseline):
            opt = baseline[opt_key]
            desc, t_ip, t_port, proto = opt.get("description", ""), opt.get("ip", ""), opt.get("port", ""), opt.get("protocol", "telnet")
            vendor = opt.get("vendor", "cisco").upper()
            if not desc:
                slabel, note, scolor = "Khong co desc", "O description trong", C_NO_DS
                cnt["no_desc"] += 1
            else:
                vr = verify_st.get((alias, opt_key))
                if vr is None:
                    slabel, note, scolor = "Chua Verify", "Chua co du lieu verify", C_UNVER
                    cnt["unverif"] += 1
                elif vr["status"] == "OK":
                    slabel, note, scolor = "OK - Khop", f"Hostname thuc te: {vr.get('act_host', '')}", C_MATCH
                    cnt["match"] += 1
                elif vr["status"] == "CANH BAO":
                    slabel, note, scolor = "SAI - Sai desc", f"Hostname thuc: {vr.get('act_host', '')}", C_WRONG
                    cnt["wrong"] += 1
                elif vr["status"] in ("TIMEOUT", "KHONG PIVOT"):
                    slabel, note, scolor = "Khong ket noi duoc", f"Trang thai: {vr['status']}", C_NO_CON
                    cnt["no_conn"] += 1
                else:
                    slabel, note, scolor = f"? {vr['status']}", vr.get("status", ""), C_UNVER
                    cnt["unverif"] += 1

            for ci, val in enumerate([ip, alias, ping_lbl, menu_state_lbl, hn, mn or "", opt_key, desc, t_ip, t_port, proto, vendor, slabel, note], 1):
                c = ws1.cell(ri, ci, val); c.border = mk_bdr(); c.alignment = Alignment(vertical="center")
                if ci == 13: c.fill = mk_fill(scolor); c.font = Font(bold=True) 
            ri += 1
            
        if ri > start_ri + 1:
            for ci in range(1, 7):
                ws1.merge_cells(start_row=start_ri, start_column=ci, end_row=ri-1, end_column=ci)
                
        summary.append({"alias": alias, "ip": ip, "ping": ping_lbl, "menu_state": menu_state_lbl, "hn": hn, "mn": mn or "", "total": len(baseline), **cnt})
    for i, w in enumerate([16, 16, 12, 20, 18, 20, 12, 38, 16, 12, 10, 12, 22, 44], 1): ws1.column_dimensions[get_column_letter(i)].width = w

    ws2 = wb.create_sheet("Tom tat")
    h2 = ["OOB Alias", "OOB IP", "Ping", "Trang thai Menu", "Hostname OOB", "Menu", "Tong Option", "OK Khop", "SAI", "Chua Verify", "Khong KN", "Khong Desc"]
    for ci, h in enumerate(h2, 1):
        c = ws2.cell(1, ci, h); c.font = Font(bold=True, color="FFFFFF"); c.fill = mk_fill(C_HDR2); c.alignment = Alignment(horizontal="center"); c.border = mk_bdr()
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(h2))}1"
    
    for ri2, sd in enumerate(summary, 2):
        for ci, val in enumerate([sd["alias"], sd["ip"], sd["ping"], sd["menu_state"], sd["hn"], sd["mn"], sd["total"], sd["match"], sd["wrong"], sd["unverif"], sd["no_conn"], sd["no_desc"]], 1):
            c = ws2.cell(ri2, ci, val); c.border = mk_bdr(); c.alignment = Alignment(vertical="center", horizontal="center" if ci>4 else "left")
        if sd.get("wrong", 0) > 0: ws2.cell(ri2, 9).fill = mk_fill(C_WRONG); ws2.cell(ri2, 9).font = Font(bold=True)
        if sd.get("match", 0) > 0: ws2.cell(ri2, 8).fill = mk_fill(C_MATCH)
    for i, w in enumerate([18, 16, 12, 20, 20, 20, 14, 10, 10, 14, 12, 13], 1): ws2.column_dimensions[get_column_letter(i)].width = w

    ws3 = wb.create_sheet("Canh bao")
    h3 = ["OOB IP", "OOB Alias", "Ping", "Trang thai Menu", "OOB Hostname", "Menu Name", "Option Key", "Description", "Target IP", "Target Port", "Protocol", "Vendor", "Desc Status", "Ghi chu"]
    for ci, h in enumerate(h3, 1):
        c = ws3.cell(1, ci, h); c.font = Font(bold=True, color="FFFFFF"); c.fill = mk_fill("C00000"); c.alignment = Alignment(horizontal="center"); c.border = mk_bdr()
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(h3))}1"

    alert_statuses = {"SAI - Sai desc", "Khong ket noi duoc"}
    ri3 = 2
    for row_idx in range(2, ri):
        if ws1.cell(row_idx, 13).value in alert_statuses: # Đã đổi 12 -> 13
            for ci in range(1, len(h3) + 1):
                dst = ws3.cell(ri3, ci, ws1.cell(row_idx, ci).value); dst.border = mk_bdr(); dst.alignment = Alignment(vertical="center")
                if ci == 13: dst.fill = mk_fill(C_WRONG); dst.font = Font(bold=True); dst.alignment = Alignment(horizontal="center", vertical="center") # Đã đổi 12 -> 13
            ri3 += 1
    if ri3 == 2: ws3.cell(2, 1, "(Khong co canh bao nao)")
    for i, w in enumerate([16, 16, 12, 20, 18, 20, 12, 38, 16, 12, 10, 12, 22, 44], 1): ws3.column_dimensions[get_column_letter(i)].width = w

    os.makedirs("reports", exist_ok=True)
    out = os.path.join("reports", f"OOB_Menu_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    wb.save(out)
    _con.print(f"\n  [green](OK)[/] Da xuat bao cao Excel: [bold]{out}[/]")

# ---------------------------------------------------------------------------
# Daemon Heartbeat
# ---------------------------------------------------------------------------
DAEMON_PID_FILE = "daemon.pid"
def _daemon_heartbeat_loop():
    while True:
        try:
            with open(DAEMON_PID_FILE, "w") as f: f.write(f"{os.getpid()}\n{datetime.now().isoformat()}\n")
        except OSError: pass
        time.sleep(30)

def _get_daemon_status() -> str:
    try:
        with open(DAEMON_PID_FILE, "r") as f: lines = f.read().splitlines()
        age = (datetime.now() - datetime.fromisoformat(lines[1])).total_seconds()
        if age < 90: return f"[green bold]RUNNING[/] [dim](PID {lines[0]})[/]"
        return f"[yellow bold]STALE[/]"
    except Exception: return "[dim]KHONG RO[/]"

def _show_menu(cfg):
    hosts_n = len(load_ip_list_cached(cfg["ip_list"]))
    user = f"[bold]{cfg['username']}[/]" if cfg["username"] else "[dim yellow](chua dat)[/]"
    menu_label = cfg['menu_name_override'] or "tu dong do"
    
    info = Text.from_markup(f"Thiet bi : [bold]{hosts_n}[/]   Menu: [bold]{menu_label}[/]   User: {user} (+{len(cfg.get('credentials', []))} phu)\nDaemon   : {_get_daemon_status()}")
    grid = Table.grid(padding=(0, 2))
    grid.add_column(style="bold cyan", min_width=5, justify="right"); grid.add_column(min_width=40)
    grid.add_row("[1]", "Them thiet bi OOB"); grid.add_row("[2]", "Xoa thiet bi OOB"); grid.add_row("\\[i]", "[cyan]Import danh sach OOB tu file Excel (.xlsx)[/]")
    grid.add_row("", ""); grid.add_row("[3]", "Cau hinh (username/password/port/multi-account...)")
    grid.add_row("[4]", "Xem danh sach thiet bi"); grid.add_row("[5]", "Xem baseline (Chuan)"); grid.add_row("[6]", "Tim kiem thiet bi")
    grid.add_row("[7]", "[bold green]Quet kiem tra Cau hinh tuc thi (Chi dinh hoac Tat ca)[/]")
    grid.add_row("[8]", "[bold magenta]Deep Verify Vat ly tuc thi (Chi dinh hoac Tat ca)[/]")
    grid.add_row("\\[p]", "[bold yellow]Push cau hinh sua loi Description (Thu cong)[/]")
    grid.add_row("[9]", "[dim magenta]Xem ket qua Verify vat ly gan nhat[/]")
    grid.add_row("\\[e]", "[yellow]Xuat bao cao menu OOB ra Excel (3 Sheet)[/]")
    grid.add_row("", ""); grid.add_row("[0]", "[bold red]Thoat[/]")

    _con.print()
    _con.print(Panel(Group(info, Rule(style="dim cyan"), grid), title="[bold cyan] [ OOB NETWORK MANAGER ] [/]", border_style="cyan", padding=(1, 2)))

def main_menu(cfg, config_path):
    while True:
        _show_menu(cfg)
        try: choice = _con.input("[bold]Chon[/]: ").strip().lower()
        except (EOFError, KeyboardInterrupt): sys.exit(0)
        _con.print()

        if choice == "1":
            ip = _con.input("  [cyan]IP thiet bi OOB[/]: ").strip()
            alias = _con.input("  [cyan]Ten goi (alias)[/]: ").strip() or None
            if ip: add_ip(cfg["ip_list"], ip, alias)
        elif choice == "2":
            ip = _con.input("  [cyan]IP can xoa[/]: ").strip()
            if ip: remove_ip(cfg["ip_list"], ip)
        elif choice == "3": settings_menu(cfg, config_path)
        elif choice == "4": list_devices(cfg)
        elif choice == "5": view_baseline(cfg)
        elif choice == "6": search_device(cfg)
        elif choice == "7": scan_specific_devices(cfg)
        elif choice == "8": verify_specific_devices(cfg)
        elif choice == "p": manual_push_devices(cfg)
        elif choice == "9": view_latest_verify_log()
        elif choice == "i": import_from_excel(cfg)
        elif choice == "e": export_menu_report(cfg)
        elif choice == "0": sys.exit(0)
        else: _con.print("  [red][!][/] Lua chon khong hop le.")

def main():
    config_path = CONFIG_FILE_DEFAULT
    for arg in sys.argv[1:]:
        if not arg.startswith("--"): config_path = arg; break
    cfg = load_config(config_path)
    if not os.path.exists(config_path): save_config(config_path, cfg)

    if "--daemon" in sys.argv: return run_daemon(cfg, config_path)
    elif "--menu" in sys.argv: return main_menu(cfg, config_path)

    _con.print(Panel("[bold cyan]1.[/] Mo Menu Quan Ly (Them/Sua IP, Xem danh sach)\n[bold cyan]2.[/] Mo Trinh Giam Sat (Chay log Daemon o terminal nay)\n[bold cyan]3.[/] Mo CA HAI (Tu dong mo 2 cua so - Yeu cau Windows)", title="[bold yellow] [ OOB LAUNCHER MULTI-TERM ] [/]", border_style="yellow", padding=(1,2)))
    choice = input("Chon che do: ").strip()
    
    if choice == "1": main_menu(cfg, config_path)
    elif choice == "2": run_daemon(cfg, config_path)
    elif choice == "3":
        if platform.system() == "Windows":
            subprocess.Popen(f'start cmd /k "python {sys.argv[0]} --daemon"', shell=True)
            subprocess.Popen(f'start cmd /k "python {sys.argv[0]} --menu"', shell=True)
        else: _con.print("Tren Linux/WSL, vui long mo 2 tab va tu chay tham so --menu / --daemon.")
        sys.exit(0)

if __name__ == "__main__":
    main()