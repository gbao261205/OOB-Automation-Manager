"""
oob_web.py - Web Panel v2.3 cho OOB Network Manager 
(Dynamic UI + Tính năng Đổi Mật Khẩu Trực Tiếp)
Chay: python oob_web.py   |   Mo trinh duyet: http://127.0.0.1:5000
"""

import json, os, queue, sqlite3, threading, time, concurrent.futures, hashlib, urllib.request
from datetime import datetime
from functools import wraps
from flask import Flask, Response, jsonify, render_template_string, request, send_file, session, redirect, url_for
from werkzeug.security import generate_password_hash, check_password_hash
import oob_monitor

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", os.urandom(24))

def _cfg(): return oob_monitor.load_config(oob_monitor.CONFIG_FILE_DEFAULT)

# --- INIT DATABASE TÀI KHOẢN ---
def _init_users_db():
    db = _cfg().get("baseline_db", "baseline.db")
    conn = sqlite3.connect(db)
    cur = conn.cursor()
    cur.execute('''CREATE TABLE IF NOT EXISTS web_users (
                    username TEXT PRIMARY KEY,
                    password_hash TEXT NOT NULL
                  )''')
    cur.execute("SELECT COUNT(*) FROM web_users")
    if cur.fetchone()[0] == 0:
        # Tự động tạo tài khoản admin/admin ở lần chạy đầu tiên
        cur.execute("INSERT INTO web_users (username, password_hash) VALUES (?, ?)", 
                    ('admin', generate_password_hash('admin')))
    conn.commit()
    conn.close()

_init_users_db()

# --- KIỂM TRA MẬT KHẨU LỘ LỌT ---
def check_pwned_password(password):
    sha1_pwd = hashlib.sha1(password.encode('utf-8')).hexdigest().upper()
    prefix, suffix = sha1_pwd[:5], sha1_pwd[5:]
    try:
        req = urllib.request.Request(f"https://api.pwnedpasswords.com/range/{prefix}", headers={'User-Agent': 'OOB-Web-Manager'})
        with urllib.request.urlopen(req, timeout=5) as res:
            hashes = (line.decode('utf-8').split(':') for line in res)
            for h, count in hashes:
                if h == suffix:
                    return int(count)
        return 0
    except Exception:
        return -1

# --- AUTH DECORATOR ---
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            if request.path.startswith('/api/'):
                return jsonify({"error": "Unauthorized"}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# --- SSE ---
_sse_clients = []
_sse_lock = threading.Lock()

def _sse_broadcast(data):
    import json as _j
    msg = "data: " + _j.dumps(data, ensure_ascii=False) + "\n\n"
    with _sse_lock:
        dead = []
        for q in _sse_clients:
            try: q.put_nowait(msg)
            except queue.Full: dead.append(q)
        for q in dead: _sse_clients.remove(q)

def _make_print_fn(task_id, ip=None):
    import re
    _re = re.compile(r"\[/?[^\[\]]*\]")
    def _fn(msg):
        clean = _re.sub("", str(msg)).strip()
        _sse_broadcast({"type":"log","task":task_id,"ip":ip,"msg":clean,"ts":datetime.now().strftime("%H:%M:%S")})
    return _fn

# --- Task Manager ---
_tasks = {}
_tasks_lock = threading.Lock()
TASKS_FILE = "task_history.json"

def _load_tasks():
    global _tasks
    try:
        if os.path.exists(TASKS_FILE):
            with open(TASKS_FILE, "r") as f: _tasks = json.load(f)
            for t in _tasks.values():
                if t.get("status") == "running": t["status"] = "error"
    except: _tasks = {}

_load_tasks()

def _save_tasks():
    try:
        with _tasks_lock:
            keys = sorted(_tasks.keys(), key=lambda k: _tasks[k].get("started", 0), reverse=True)[:50]
            to_save = {k: _tasks[k] for k in keys}
        with open(TASKS_FILE, "w") as f: json.dump(to_save, f)
    except: pass

def _new_task(tid, action="unknown", ip=None):
    with _tasks_lock: _tasks[tid] = {"status":"running","started":time.time(),"action":action,"ip":ip}
    _save_tasks()

def _finish_task(tid, error=None):
    with _tasks_lock:
        if tid in _tasks:
            _tasks[tid].update({
                "status": "error" if error else "done",
                "finished": time.time(),
                "error": str(error) if error else None
            })
    _save_tasks()
    _sse_broadcast({"type":"task_done","task":tid,"status":"error" if error else "done"})


def _query_bl(q, args=(), all_rows=True):
    db = _cfg().get("baseline_db","baseline.db")
    if not os.path.exists(db): return [] if all_rows else None
    try:
        conn = sqlite3.connect(db); conn.row_factory = sqlite3.Row
        cur = conn.cursor(); cur.execute(q, args)
        rv = cur.fetchall() if all_rows else cur.fetchone(); conn.close(); return rv
    except: return [] if all_rows else None

# --- Background runners ---
def _run_scan(tid, target_ip=None):
    try:
        cfg = _cfg(); hosts = oob_monitor.load_ip_list_cached(cfg["ip_list"])
        if target_ip: hosts = [h for h in hosts if h[0] == target_ip]
        pfn = _make_print_fn(tid, target_ip)
        pfn("Bat dau SCAN " + str(len(hosts)) + " thiet bi (5 threads)...")
        def _scan_single(h):
            ip, alias = h[0], h[1]
            pfn("  [PING] " + alias + " (" + ip + ")")
            alive = oob_monitor.ping_host(ip)
            oob_monitor.save_device_status(ip, alias=alias, ping=alive)
            if not alive: pfn("  [!] " + alias + ": Khong ping duoc."); return
            pfn("  [SCAN] " + alias + " (" + ip + ")")
            try: hn, mn, snap, ms = oob_monitor.poll_host_multi(ip, cfg, timeout=10)
            except Exception as e:
                oob_monitor.save_device_status(ip, alias=alias, menu_state="conn_failed")
                pfn("  [LOI] " + alias + ": " + str(e)); return
            oob_monitor.save_device_status(ip, alias=alias, menu_state=ms)
            if ms in ["fetch_failed","no_menu"] or not snap: pfn("  [!] " + alias + ": Khong co menu."); return
            oob_monitor.save_options(cfg["snapshot_db"],"snapshot_menu",ip,mn,hn,snap)
            _,_,bl = oob_monitor.get_options_by_host(cfg["baseline_db"],"baseline_menu",ip)
            if bl is None or not oob_monitor.options_equal(bl, snap):
                oob_monitor.save_options(cfg["baseline_db"],"baseline_menu",ip,mn,hn,snap)
                oob_monitor.log_baseline_change(alias, ip, "CAP NHAT QUA WEB")
                pfn("  [OK] " + alias + ": Cap nhat baseline (" + str(len(snap)) + " option).")
            else: pfn("  [OK] " + alias + ": Khop baseline.")
            
        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as ex:
            ex.map(_scan_single, hosts)
        pfn("[OK] Hoan thanh SCAN!"); _finish_task(tid)
    except Exception as e: _finish_task(tid, e)

def _run_verify(tid, target_ip=None):
    try:
        cfg = _cfg(); hosts = oob_monitor.load_ip_list_cached(cfg["ip_list"])
        if target_ip: hosts = [h for h in hosts if h[0] == target_ip]
        pfn = _make_print_fn(tid, target_ip)
        pfn("Bat dau VERIFY " + str(len(hosts)) + " thiet bi (5 threads)...")
        def _verify_single(h):
            ip, alias = h[0], h[1]
            _,_,bl = oob_monitor.get_options_by_host(cfg["baseline_db"],"baseline_menu",ip)
            if not bl: pfn("  [!] " + alias + ": Chua co baseline."); return
            pfn("  [VERIFY] " + alias + " (" + ip + ")")
            oob_monitor.run_deep_verify(cfg, alias, ip, bl, print_fn=pfn)
            
        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as ex:
            ex.map(_verify_single, hosts)
        pfn("[OK] Hoan thanh VERIFY!"); _finish_task(tid)
    except Exception as e: _finish_task(tid, e)

def _run_push(tid, target_ip=None):
    try:
        cfg = _cfg(); hosts = oob_monitor.load_ip_list_cached(cfg["ip_list"])
        if target_ip: hosts = [h for h in hosts if h[0] == target_ip]
        pfn = _make_print_fn(tid, target_ip)
        pfn("Bat dau PUSH " + str(len(hosts)) + " thiet bi (5 threads)...")
        def _push_single(h):
            ip, alias = h[0], h[1]
            _,_,bl = oob_monitor.get_options_by_host(cfg["baseline_db"],"baseline_menu",ip)
            if not bl: return
            vendor = next(iter(bl.values())).get("vendor","cisco") if bl else "cisco"
            if vendor == "vertiv": pfn("  [!] " + alias + ": Vertiv khong ho tro Push."); return
            results = oob_monitor.run_deep_verify(cfg, alias, ip, bl, print_fn=pfn)
            if any(r["status"]=="CANH BAO" for r in results):
                oob_monitor.process_push_and_reverify(cfg, alias, ip, bl, results, print_fn=pfn)
            else: pfn("  [OK] " + alias + ": Khong co sai lech.")
            
        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as ex:
            ex.map(_push_single, hosts)
        pfn("[OK] Hoan thanh PUSH!"); _finish_task(tid)
    except Exception as e: _finish_task(tid, e)

# --- WEB & API ROUTES ---
LOGIN_HTML = r"""<!DOCTYPE html>
<html lang="vi">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Đăng nhập - OOB Manager</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
body { background: #07070f; color: #f1f0ff; font-family: 'Inter', sans-serif; display: flex; height: 100vh; align-items: center; justify-content: center; margin: 0; }
.login-box { background: rgba(255,255,255,0.04); border: 1px solid rgba(255,255,255,0.1); padding: 35px 30px; border-radius: 12px; width: 340px; box-shadow: 0 10px 40px rgba(0,0,0,0.5); }
.title { font-size: 20px; font-weight: 600; margin-bottom: 25px; text-align: center; }
.fg { margin-bottom: 15px; }
.fc { width: 100%; background: rgba(255,255,255,0.06); border: 1px solid rgba(255,255,255,0.1); color: #fff; padding: 12px; border-radius: 6px; outline: none; box-sizing: border-box; font-family: inherit;}
.fc:focus { border-color: #7c3aed; box-shadow: 0 0 0 3px rgba(124,58,237,.15); }
.btn { width: 100%; background: #7c3aed; color: #fff; border: none; padding: 12px; border-radius: 6px; font-weight: 600; cursor: pointer; transition: 0.2s; margin-top: 10px; font-size: 14px;}
.btn:hover { background: #6d28d9; }
.btn-g { background: rgba(255,255,255,0.1); margin-top: 10px; }
.btn-g:hover { background: rgba(255,255,255,0.2); }
.msg { font-size: 13px; margin-top: 15px; text-align: center; display: none; line-height: 1.4; padding: 10px; border-radius: 6px; font-weight: 500;}
.msg.err { color: #ef4444; background: rgba(239, 68, 68, 0.1); border: 1px solid rgba(239, 68, 68, 0.2); }
.msg.warn { color: #fbbf24; background: rgba(251, 191, 36, 0.1); border: 1px solid rgba(251, 191, 36, 0.2); }
.msg.ok { color: #22c55e; background: rgba(34, 197, 94, 0.1); border: 1px solid rgba(34, 197, 94, 0.2); }
</style>
</head>
<body>
<div class="login-box">
    <div class="title">🔑 OOB Web Manager</div>
    <div class="fg"><input type="text" id="usr" class="fc" placeholder="Tên đăng nhập"></div>
    <div class="fg"><input type="password" id="pwd" class="fc" placeholder="Mật khẩu"></div>
    <button class="btn" onclick="doLogin()">Đăng nhập Hệ thống</button>
    <button class="btn btn-g" onclick="window.location.href='/'">Quay lại Tra cứu</button>
    <div id="msg" class="msg"></div>
</div>
<script>
async function doLogin() {
    const u = document.getElementById('usr').value.trim(), p = document.getElementById('pwd').value;
    if(!u || !p) return;
    const m = document.getElementById('msg'); m.style.display = 'block'; m.className = 'msg'; m.textContent = 'Đang xác thực...';
    
    try {
        const r = await fetch('/login', { method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({username:u, password:p}) });
        const d = await r.json();
        
        if (d.status === 'error') { m.className = 'msg err'; m.textContent = d.msg; }
        else if (d.status === 'warning') { m.className = 'msg warn'; m.textContent = d.msg; setTimeout(()=>window.location.href=d.redirect, 3500); }
        else { m.className = 'msg ok'; m.textContent = d.msg; setTimeout(()=>window.location.href=d.redirect, 1000); }
    } catch(e) { m.className = 'msg err'; m.textContent = 'Mất kết nối tới server.'; }
}
document.addEventListener('keydown', e => { if(e.key === 'Enter') doLogin(); });
</script>
</body>
</html>"""

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        data = request.json
        username = data.get('username', '').strip()
        password = data.get('password', '')
        
        db = _cfg().get("baseline_db", "baseline.db")
        conn = sqlite3.connect(db)
        cur = conn.cursor()
        cur.execute("SELECT password_hash FROM web_users WHERE username=?", (username,))
        row = cur.fetchone()
        conn.close()
        
        if row and check_password_hash(row[0], password):
            session['logged_in'] = True
            session['username'] = username
            
            leak_count = check_pwned_password(password)
            if leak_count > 0:
                return jsonify({
                    "status": "warning", 
                    "msg": f"Xác thực thành công. CẢNH BÁO: Mật khẩu này đã bị lộ {leak_count} lần trên mạng! Khuyến nghị đổi mật khẩu.", 
                    "redirect": "/"
                })
            elif leak_count == -1:
                return jsonify({"status": "ok", "msg": "Xác thực thành công. (Chưa thể kết nối API kiểm tra rò rỉ)", "redirect": "/"})
                
            return jsonify({"status": "ok", "msg": "Xác thực thành công! Mật khẩu an toàn.", "redirect": "/"})
        
        return jsonify({"status": "error", "msg": "Sai tài khoản hoặc mật khẩu."}), 401
        
    return render_template_string(LOGIN_HTML)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route("/api/change-password", methods=["POST"])
@login_required
def api_change_password():
    data = request.json or {}
    old_password = data.get("old_password", "")
    new_password = data.get("new_password", "")
    username = session.get("username")

    if not old_password or not new_password:
        return jsonify({"status": "error", "msg": "Vui lòng nhập đầy đủ thông tin mật khẩu!"}), 400

    db = _cfg().get("baseline_db", "baseline.db")
    conn = sqlite3.connect(db)
    cur = conn.cursor()
    cur.execute("SELECT password_hash FROM web_users WHERE username=?", (username,))
    row = cur.fetchone()

    # Kiểm tra mật khẩu cũ
    if not row or not check_password_hash(row[0], old_password):
        conn.close()
        return jsonify({"status": "error", "msg": "Mật khẩu cũ không chính xác!"})

    # Cập nhật mật khẩu mới
    new_hash = generate_password_hash(new_password)
    cur.execute("UPDATE web_users SET password_hash=? WHERE username=?", (new_hash, username))
    conn.commit()
    conn.close()

    # Kiểm tra lộ lọt cho mật khẩu mới
    leak_count = check_pwned_password(new_password)
    warn_msg = ""
    if leak_count > 0:
        warn_msg = f" (Lưu ý: Mật khẩu mới đã từng bị lộ {leak_count} lần trên mạng toàn cầu, hãy cẩn thận!)"

    return jsonify({"status": "ok", "msg": f"Đổi mật khẩu thành công!{warn_msg}"})

# CÁC API KHÔNG CẦN ĐĂNG NHẬP (Read-only / Tra cứu)
@app.route("/api/events")
def api_events():
    q = queue.Queue(maxsize=100)
    with _sse_lock: _sse_clients.append(q)
    def gen():
        try:
            yield 'data: {"type":"connected"}\n\n'
            while True:
                try: yield q.get(timeout=25)
                except queue.Empty: yield ": hb\n\n"
        except GeneratorExit:
            with _sse_lock:
                if q in _sse_clients: _sse_clients.remove(q)
    return Response(gen(), mimetype="text/event-stream", headers={"Cache-Control":"no-cache","X-Accel-Buffering":"no"})


@app.route("/api/stats")
def api_stats():
    cfg = _cfg(); ips = oob_monitor.load_ip_list_cached(cfg["ip_list"])
    ds = oob_monitor.load_device_status()
    stats = {"total":len(ips),"online":0,"offline":0,"has_baseline":0,"alarms":0}
    for ip,_ in ips:
        st = ds.get(ip,{})
        if st.get("ping") is True: stats["online"] += 1
        elif st.get("ping") is False: stats["offline"] += 1
        r = _query_bl("SELECT COUNT(*) as c FROM baseline_menu WHERE host=?",(ip,),all_rows=False)
        if r and r["c"] > 0: stats["has_baseline"] += 1
    vst = oob_monitor._parse_verify_logs_for_status(max_age_hours=24.0)
    for v in vst.values():
        if v.get("status") == "CANH BAO": stats["alarms"] += 1
    return jsonify(stats)

@app.route("/api/devices")
def api_devices():
    cfg = _cfg(); ips = oob_monitor.load_ip_list_cached(cfg["ip_list"])
    ds = oob_monitor.load_device_status()
    vst = oob_monitor._parse_verify_logs_for_status(max_age_hours=24.0*7)
    devs = []
    for ip, alias in ips:
        st = ds.get(ip,{})
        r = _query_bl("SELECT COUNT(*) as c FROM baseline_menu WHERE host=?",(ip,),all_rows=False)
        opt_count = r["c"] if r else 0
        alarm_c = sum(1 for (a,k),v in vst.items() if a==alias and v.get("status")=="CANH BAO")
        ok_c = sum(1 for (a,k),v in vst.items() if a==alias and v.get("status")=="OK")
        mn, dn, _ = oob_monitor.get_options_by_host(cfg["baseline_db"],"baseline_menu",ip)
        upd = oob_monitor.get_updated_at_by_host(cfg["baseline_db"],"baseline_menu",ip)
        devs.append({"ip":ip,"alias":alias,"ping":st.get("ping"),"menu_state":st.get("menu_state"),
            "checked_at":st.get("checked_at","-"),"opt_count":opt_count,
            "device_name":dn or "","menu_name":mn or "","updated_at":upd or "",
            "alarm_count":alarm_c,"ok_count":ok_c})
    return jsonify(devs)

@app.route("/api/device/<ip>/options")
def api_device_options(ip):
    cfg = _cfg(); vst = oob_monitor._parse_verify_logs_for_status(max_age_hours=24.0*7)
    mn, dn, bl = oob_monitor.get_options_by_host(cfg["baseline_db"],"baseline_menu",ip)
    if not bl: return jsonify({"device_name":dn,"menu_name":mn,"options":[]})
    all_ips = oob_monitor.load_ip_list_cached(cfg["ip_list"])
    alias = next((a for i,a in all_ips if i==ip), ip)
    opts = []
    for key in sorted(bl.keys()):
        o = bl[key]; vr = vst.get((alias,key))
        opts.append({"key":key,"description":o.get("description",""),"ip":o.get("ip",""),
            "port":o.get("port",23),"protocol":o.get("protocol","telnet"),"vendor":o.get("vendor","cisco"),
            "verify_status":vr["status"] if vr else None,"act_host":vr.get("act_host") if vr else None})
    return jsonify({"device_name":dn,"menu_name":mn,"alias":alias,"options":opts})

@app.route("/api/tasks")
def api_tasks():
    with _tasks_lock: return jsonify(dict(_tasks))

@app.route("/api/search")
def api_search():
    query = request.args.get("q","").strip().lower()
    if not query: return jsonify([])
    cfg = _cfg(); hosts = oob_monitor.load_ip_list_cached(cfg["ip_list"])
    vst = oob_monitor._parse_verify_logs_for_status(max_age_hours=24.0*30)
    found = []
    for ip, alias in hosts:
        _mn,dn,source = oob_monitor.get_options_by_host(cfg["baseline_db"],"baseline_menu",ip)
        if source is None: _mn,dn,source = oob_monitor.get_options_by_host(cfg["snapshot_db"],"snapshot_menu",ip)
        if not source: continue
        dn = dn or alias
        for key, entry in source.items():
            ah = vst.get((alias,key),{}).get("act_host","") or ""
            sc = 0
            if query==ah.lower(): sc=100
            elif query in ah.lower(): sc=90
            elif query==alias.lower() or query==ip: sc=85
            elif query==dn.lower(): sc=80
            elif query in alias.lower() or query in dn.lower(): sc=75
            elif query in entry.get("description","").lower(): sc=60
            elif query in entry.get("ip","").lower(): sc=50
            elif query==key.lower(): sc=40
            if sc>0:
                found.append({"score":sc,"oob_ip":ip,"oob_alias":alias,"oob_host":dn,
                    "opt_key":key,"desc":entry.get("description",""),"target_ip":entry.get("ip",""),
                    "target_port":entry.get("port",23),"protocol":entry.get("protocol","telnet"),
                    "act_host":ah,"verify_status":vst.get((alias,key),{}).get("status")})
    found.sort(key=lambda x:x["score"],reverse=True)
    return jsonify(found[:100])

@app.route("/api/logs")
def api_logs():
    d = "verify-logs"
    if not os.path.exists(d): return jsonify([])
    files = [{"name":f,"size":os.path.getsize(os.path.join(d,f)),
              "mtime":datetime.fromtimestamp(os.path.getmtime(os.path.join(d,f))).strftime("%Y-%m-%d %H:%M:%S")}
             for f in sorted(os.listdir(d),reverse=True) if f.endswith(".log")]
    return jsonify(files[:50])

@app.route("/api/logs/<path:fn>")
def api_log_content(fn):
    fp = os.path.join("verify-logs", os.path.basename(fn))
    if not os.path.exists(fp): return jsonify({"error":"Not found"}),404
    try:
        with open(fp,"r",encoding="utf-8") as f: return jsonify({"content":f.read()})
    except Exception as e: return jsonify({"error":str(e)}),500

@app.route("/api/push-logs/<path:fn>")
def api_push_log_content(fn):
    fp = os.path.join("push-logs", os.path.basename(fn))
    if not os.path.exists(fp): return jsonify({"error":"Not found"}),404
    try:
        with open(fp,"r",encoding="utf-8") as f: return jsonify({"content":f.read()})
    except Exception as e: return jsonify({"error":str(e)}),500

@app.route("/api/push-logs")
def api_push_logs():
    d = "push-logs"
    if not os.path.exists(d): return jsonify([])
    files = [{"name":f,"size":os.path.getsize(os.path.join(d,f)),
              "mtime":datetime.fromtimestamp(os.path.getmtime(os.path.join(d,f))).strftime("%Y-%m-%d %H:%M:%S")}
             for f in sorted(os.listdir(d),reverse=True) if f.endswith(".log")]
    return jsonify(files[:50])

@app.route("/api/export/excel")
def api_export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError: return jsonify({"error":"Thieu openpyxl. pip install openpyxl"}),500
    cfg = _cfg(); hosts = oob_monitor.load_ip_list_cached(cfg["ip_list"])
    vst = oob_monitor._parse_verify_logs_for_status(); ds = oob_monitor.load_device_status()
    def mk_fill(c): return PatternFill(fill_type="solid",fgColor=c)
    def mk_bdr():
        s = Side(style="thin",color="BFBFBF"); return Border(left=s,right=s,top=s,bottom=s)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Chi tiet OOB"
    hdrs = ["OOB IP","OOB Alias","Hostname","Menu","Ping","Menu State","Option Key","Description","Target IP","Port","Protocol","Verify","Act Host"]
    for ci,h in enumerate(hdrs,1):
        c = ws.cell(1,ci,h); c.font=Font(bold=True,color="FFFFFF"); c.fill=mk_fill("1F4E79"); c.alignment=Alignment(horizontal="center"); c.border=mk_bdr()
    ws.freeze_panes = "A2"; ri = 2
    for ip,alias in hosts:
        st = ds.get(ip,{})
        pg = "Online" if st.get("ping") is True else ("Offline" if st.get("ping") is False else "-")
        mn,dn,bl = oob_monitor.get_options_by_host(cfg["baseline_db"],"baseline_menu",ip)
        if not bl:
            for ci,v in enumerate([ip,alias,"","",pg,st.get("menu_state","-"),"(chua co baseline)","","","","","",""],1): ws.cell(ri,ci,v).border=mk_bdr()
            ri+=1; continue
        for ok_key in sorted(bl):
            o = bl[ok_key]; vr = vst.get((alias,ok_key))
            if vr:
                vs = vr["status"]; ah = vr.get("act_host","") or ""
                sc = "C6EFCE" if vs=="OK" else ("FFC7CE" if vs=="CANH BAO" else "FFCC99")
            else: vs,ah,sc = "Chua Verify","","FFEB9C"
            for ci,v in enumerate([ip,alias,dn or "",mn or "",pg,st.get("menu_state","-"),ok_key,o.get("description",""),o.get("ip",""),o.get("port",""),o.get("protocol",""),vs,ah],1):
                c = ws.cell(ri,ci,v); c.border=mk_bdr()
                if ci==12: c.fill=mk_fill(sc); c.font=Font(bold=True)
            ri+=1
    for i,w in enumerate([16,16,18,18,10,14,12,36,16,8,10,14,18],1): ws.column_dimensions[get_column_letter(i)].width=w
    os.makedirs("reports",exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fn = "OOB_Report_" + ts + ".xlsx"; fp = os.path.join("reports",fn)
    wb.save(fp)
    return send_file(fp,as_attachment=True,download_name=fn,mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# CÁC API KHÓA HÀNH ĐỘNG THAY ĐỔI DỮ LIỆU (Action / Quản trị)
@app.route("/api/config", methods=["GET","POST"])
@login_required
def api_config():
    if request.method == "GET":
        return jsonify({k:v for k,v in oob_monitor.load_config(oob_monitor.CONFIG_FILE_DEFAULT).items() if k!="credentials"})
    cfg = oob_monitor.load_config(oob_monitor.CONFIG_FILE_DEFAULT)
    allowed = ["username","password","enable_password","vertiv_connect_password",
               "menu_name_override","ssh_port","telnet_port","interval","verify_interval",
               "ip_list","baseline_db","snapshot_db","auto_verify","verify_schedule_mode",
               "verify_schedule_time","verify_schedule_weekday","scan_schedule_mode",
               "scan_schedule_time","scan_schedule_weekday","verify_wait_after_connect","max_verify_duration"]
    for k in allowed:
        if k in (request.json or {}): cfg[k] = request.json[k]
    oob_monitor.save_config(oob_monitor.CONFIG_FILE_DEFAULT, cfg)
    return jsonify({"status":"ok"})

@app.route("/api/credentials", methods=["GET","POST","DELETE"])
@login_required
def api_creds():
    cfg = oob_monitor.load_config(oob_monitor.CONFIG_FILE_DEFAULT); creds = cfg.get("credentials",[])
    if request.method == "GET":
        return jsonify([{"username":c.get("username",""),"has_pass":bool(c.get("password")),"has_enable":bool(c.get("enable_password"))} for c in creds])
    if request.method == "POST":
        d = request.json or {}
        creds.append({"username":d.get("username",""),"password":d.get("password",""),"enable_password":d.get("enable_password","")})
        cfg["credentials"] = creds; oob_monitor.save_config(oob_monitor.CONFIG_FILE_DEFAULT, cfg)
        return jsonify({"status":"ok"})
    idx = (request.json or {}).get("index",-1)
    if 0 <= idx < len(creds): creds.pop(idx)
    cfg["credentials"] = creds; oob_monitor.save_config(oob_monitor.CONFIG_FILE_DEFAULT, cfg)
    return jsonify({"status":"ok"})

@app.route("/api/device", methods=["POST","DELETE"])
@login_required
def api_device():
    cfg = _cfg()
    if request.method == "POST":
        d = request.json or {}; ip = d.get("ip","").strip(); alias = d.get("alias","").strip() or None
        if not ip: return jsonify({"status":"error"}),400
        oob_monitor.add_ip(cfg["ip_list"],ip,alias); return jsonify({"status":"ok"})
    ip = (request.json or {}).get("ip","").strip()
    oob_monitor.remove_ip(cfg["ip_list"],ip); return jsonify({"status":"ok"})

@app.route("/api/action", methods=["POST"])
@login_required
def api_action():
    d = request.json or {}; action = d.get("action"); tip = d.get("ip") or None
    tid = action + "_" + (tip or "all") + "_" + str(int(time.time()))
    _new_task(tid, action=action, ip=tip)
    runners = {"scan":_run_scan,"verify":_run_verify,"push":_run_push}
    if action not in runners: return jsonify({"status":"error"}),400
    threading.Thread(target=runners[action], args=(tid,tip), daemon=True).start()
    return jsonify({"status":"ok","task_id":tid,"msg":"Da dua lenh " + action.upper() + " vao hang doi!"})

@app.route("/api/revert", methods=["POST"])
@login_required
def api_revert():
    fn = (request.json or {}).get("filename")
    if not fn: return jsonify({"error":"Missing filename"}), 400
    fp = os.path.join("push-logs", os.path.basename(fn))
    if not os.path.exists(fp): return jsonify({"error":"Not found"}), 404
    
    import re
    revert_cmds = []
    oob_ip = None
    try:
        with open(fp, "r", encoding="utf-8") as f:
            for line in f:
                if "=== PUSH LOG TỰ ĐỘNG:" in line:
                    m = re.search(r'\(([\d\.]+)\)', line)
                    if m: oob_ip = m.group(1)
                elif "REVERT CMD:" in line:
                    m = re.search(r'menu\s+(\S+)\s+text\s+(\S+)\s+(.+)', line)
                    if m: revert_cmds.append((m.group(1), m.group(2), m.group(3).strip()))
    except Exception as e: return jsonify({"error":str(e)}), 500
    
    if not oob_ip or not revert_cmds: return jsonify({"error":"Khong the parse revert commands hoac OOB IP tu log"}), 400
    
    cfg = _cfg()
    c = oob_monitor.get_all_credentials(cfg)[0] if oob_monitor.get_all_credentials(cfg) else {"username":"","password":"","enable_password":""}
    
    def _run_rev():
        tid = "revert_" + str(int(time.time()))
        _new_task(tid, "revert", oob_ip)
        pfn = _make_print_fn(tid, oob_ip)
        pfn(f"Bat dau REVERT cho {oob_ip} ({len(revert_cmds)} options)...")
        try:
            oob_monitor.push_menu_descriptions(oob_ip, cfg.get("ssh_port",22), cfg.get("telnet_port",23), c["username"], c["password"], c["enable_password"], revert_cmds, timeout=10)
            pfn("[OK] Hoan thien Revert!")
            _finish_task(tid)
        except Exception as e:
            pfn(f"[LOI] Revert that bai: {e}")
            _finish_task(tid, e)
            
    threading.Thread(target=_run_rev, daemon=True).start()
    return jsonify({"status":"ok","msg":f"Da bat dau revert {len(revert_cmds)} muc."})

@app.route("/api/import", methods=["POST"])
@login_required
def api_import():
    import re; _IP = re.compile(r"^\d{1,3}(\.\d{1,3}){3}$")
    cfg = _cfg(); lines = (request.json or {}).get("text","").splitlines()
    added = skipped = 0; existing = {h[0] for h in oob_monitor.load_ip_list_cached(cfg["ip_list"])}
    for line in lines:
        line = line.strip()
        if not line or line.startswith("#"): continue
        parts = line.split(); ip = parts[0]; alias = parts[1] if len(parts)>1 else ip
        if not _IP.match(ip) or ip in existing: skipped+=1; continue
        oob_monitor.add_ip(cfg["ip_list"],ip,alias); existing.add(ip); added+=1
    return jsonify({"added":added,"skipped":skipped})


HTML = r"""<!DOCTYPE html>
<html lang="vi">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>OOB Network Manager</title>
<meta name="description" content="OOB Network Manager - Giam sat va quan ly thiet bi Out-of-Band">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
:root{
  --bg:#07070f;--bg-card:rgba(255,255,255,0.04);--bg-hover:rgba(255,255,255,0.08);--bg-input:rgba(255,255,255,0.06);
  --border:rgba(255,255,255,0.1);--border-hv:rgba(255,255,255,0.2);
  --violet:#7c3aed;--teal:#06d6a0;--pink:#f72585;--amber:#fbbf24;--blue:#3b82f6;--red:#ef4444;--green:#22c55e;
  --text:#f1f0ff;--text2:#a09db8;--text3:#6b6880;
  --sw:240px;--r:12px;--rs:8px;--tr:all 0.2s cubic-bezier(0.4,0,0.2,1);
}
html,body{height:100%;font-family:'Inter',sans-serif;background:var(--bg);color:var(--text);overflow-x:hidden}
::-webkit-scrollbar{width:6px;height:6px}
::-webkit-scrollbar-thumb{background:rgba(124,58,237,0.4);border-radius:3px}
::-webkit-scrollbar-thumb:hover{background:var(--violet)}
.app{display:flex;height:100vh}

/* SIDEBAR */
.sidebar{width:var(--sw);min-height:100vh;background:linear-gradient(180deg,rgba(124,58,237,0.12) 0%,rgba(7,7,15,0.95) 60%);border-right:1px solid var(--border);display:flex;flex-direction:column;position:fixed;top:0;left:0;z-index:100;backdrop-filter:blur(20px)}
.sb-logo{padding:20px;border-bottom:1px solid var(--border);display:flex;align-items:center;gap:10px}
.sb-icon{width:38px;height:38px;background:linear-gradient(135deg,var(--violet),var(--teal));border-radius:10px;display:flex;align-items:center;justify-content:center;font-size:18px;flex-shrink:0}
.sb-logo-t{font-size:14px;font-weight:700;line-height:1.2}
.sb-logo-s{font-size:10px;color:var(--text3);font-weight:400}
.sb-nav{flex:1;padding:12px 0;overflow-y:auto}
.sb-sec{padding:0 12px;margin-bottom:4px}
.sb-sec-lbl{font-size:10px;font-weight:600;color:var(--text3);text-transform:uppercase;letter-spacing:.08em;padding:8px 8px 4px}
.nav-item{display:flex;align-items:center;gap:10px;padding:9px 12px;border-radius:var(--rs);cursor:pointer;transition:var(--tr);color:var(--text2);font-size:13.5px;font-weight:500;position:relative;user-select:none;border:none;background:none;width:100%;text-align:left}
.nav-item:hover{background:var(--bg-hover);color:var(--text)}
.nav-item.active{background:linear-gradient(135deg,rgba(124,58,237,0.25),rgba(6,214,160,0.1));color:#fff;box-shadow:inset 0 0 0 1px rgba(124,58,237,0.3)}
.nav-item.active::before{content:'';position:absolute;left:0;top:50%;transform:translateY(-50%);width:3px;height:60%;background:var(--violet);border-radius:0 3px 3px 0}
.nav-ic{width:18px;height:18px;flex-shrink:0;opacity:.8}
.nav-item.active .nav-ic{opacity:1}
.nav-badge{margin-left:auto;background:var(--pink);color:#fff;font-size:10px;font-weight:700;padding:2px 7px;border-radius:100px;min-width:20px;text-align:center}
.sb-foot{padding:14px 16px;border-top:1px solid var(--border);font-size:11px;color:var(--text3)}
.d-dot{width:7px;height:7px;border-radius:50%;background:var(--text3);flex-shrink:0}
.d-dot.on{background:var(--green);box-shadow:0 0 8px var(--green);animation:pdot 2s infinite}
@keyframes pdot{0%,100%{opacity:1}50%{opacity:.4}}

/* MAIN */
.main{margin-left:var(--sw);flex:1;min-height:100vh;display:flex;flex-direction:column;overflow:hidden}
.topbar{height:60px;padding:0 24px;display:flex;align-items:center;gap:16px;background:rgba(7,7,15,.8);backdrop-filter:blur(20px);border-bottom:1px solid var(--border);position:sticky;top:0;z-index:50}
.tb-title{font-size:17px;font-weight:600;flex:1}
.sw{position:relative}
.sw input{background:var(--bg-input);border:1px solid var(--border);color:var(--text);border-radius:var(--rs);padding:7px 14px 7px 36px;font-size:13px;font-family:inherit;width:260px;transition:var(--tr);outline:none}
.sw input:focus{border-color:var(--violet);box-shadow:0 0 0 3px rgba(124,58,237,.2)}
.sw .si{position:absolute;left:11px;top:50%;transform:translateY(-50%);opacity:.5;pointer-events:none}
.content{flex:1;overflow-y:auto;padding:24px}

/* PAGES */
.page{display:none}
.page.active{display:block;animation:fIn .25s ease}
@keyframes fIn{from{opacity:0;transform:translateY(8px)}to{opacity:1;transform:translateY(0)}}

/* STATS */
.sg{display:grid;grid-template-columns:repeat(4,1fr);gap:16px;margin-bottom:24px}
.sc{background:var(--bg-card);border:1px solid var(--border);border-radius:var(--r);padding:20px;position:relative;overflow:hidden;transition:var(--tr)}
.sc:hover{border-color:var(--border-hv);transform:translateY(-2px);box-shadow:0 4px 24px rgba(0,0,0,.4)}
.sc::before{content:'';position:absolute;top:0;left:0;right:0;height:2px;background:var(--acc,var(--violet))}
.sc .sv{font-size:32px;font-weight:700;line-height:1;margin-bottom:6px}
.sc .sl{font-size:12px;color:var(--text2);font-weight:500;text-transform:uppercase;letter-spacing:.06em}
.sc .si2{position:absolute;right:16px;top:16px;font-size:28px;opacity:.12}
.sc.c1{--acc:var(--violet)}.sc.c2{--acc:var(--teal)}.sc.c3{--acc:var(--amber)}.sc.c4{--acc:var(--pink)}

/* SECTION HEADER */
.sh{display:flex;align-items:center;justify-content:space-between;margin-bottom:16px}
.st{font-size:15px;font-weight:600;display:flex;align-items:center;gap:8px}
.st .dot{width:8px;height:8px;border-radius:50%;background:var(--violet)}

/* TABLE */
.tw{background:var(--bg-card);border:1px solid var(--border);border-radius:var(--r);overflow:hidden}
table{width:100%;border-collapse:collapse}
thead tr{background:rgba(255,255,255,.04)}
th{padding:11px 14px;text-align:left;font-size:11px;font-weight:600;color:var(--text3);text-transform:uppercase;letter-spacing:.07em;border-bottom:1px solid var(--border);white-space:nowrap}
td{padding:11px 14px;border-bottom:1px solid rgba(255,255,255,.04);font-size:13px;vertical-align:middle}
tbody tr:hover{background:var(--bg-hover)}
tbody tr:last-child td{border-bottom:none}
.mono{font-family:'JetBrains Mono',monospace;font-size:12px}

/* BADGES */
.badge{display:inline-flex;align-items:center;gap:4px;padding:3px 9px;border-radius:100px;font-size:11px;font-weight:600;white-space:nowrap}
.bg{background:rgba(34,197,94,.15);color:var(--green)}
.br{background:rgba(239,68,68,.15);color:var(--red)}
.ba{background:rgba(251,191,36,.15);color:var(--amber)}
.bv{background:rgba(124,58,237,.15);color:#a78bfa}
.bt{background:rgba(6,214,160,.15);color:var(--teal)}
.bm{background:rgba(255,255,255,.08);color:var(--text3)}
.bp{background:rgba(247,37,133,.15);color:var(--pink)}
.bpulse{animation:bbl 1.5s ease infinite}
@keyframes bbl{0%,100%{opacity:1}50%{opacity:.6}}

/* BUTTONS */
.btn{display:inline-flex;align-items:center;gap:6px;padding:8px 16px;border-radius:var(--rs);font-size:13px;font-weight:500;font-family:inherit;cursor:pointer;border:none;transition:var(--tr);text-decoration:none;white-space:nowrap}
.btn:disabled{opacity:.5;cursor:not-allowed}
.btn-p{background:var(--violet);color:#fff}
.btn-p:hover:not(:disabled){background:#6d28d9;box-shadow:0 4px 15px rgba(124,58,237,.4)}
.btn-t{background:rgba(6,214,160,.15);color:var(--teal);border:1px solid rgba(6,214,160,.3)}
.btn-t:hover:not(:disabled){background:rgba(6,214,160,.25)}
.btn-a{background:rgba(251,191,36,.15);color:var(--amber);border:1px solid rgba(251,191,36,.3)}
.btn-a:hover:not(:disabled){background:rgba(251,191,36,.25)}
.btn-pk{background:rgba(247,37,133,.15);color:var(--pink);border:1px solid rgba(247,37,133,.3)}
.btn-pk:hover:not(:disabled){background:rgba(247,37,133,.25)}
.btn-g{background:var(--bg-input);color:var(--text2);border:1px solid var(--border)}
.btn-g:hover:not(:disabled){border-color:var(--border-hv);color:var(--text)}
.btn-d{background:rgba(239,68,68,.15);color:var(--red);border:1px solid rgba(239,68,68,.3)}
.btn-d:hover:not(:disabled){background:rgba(239,68,68,.25)}
.btn-sm{padding:5px 10px;font-size:12px}
.btn-ic{padding:6px;width:30px;height:30px}
.bg2{display:flex;gap:6px;flex-wrap:wrap}

/* FORMS */
.fg{margin-bottom:16px}
.fl{display:block;font-size:12px;font-weight:500;color:var(--text2);margin-bottom:6px;text-transform:uppercase;letter-spacing:.05em}
.fc{width:100%;background:var(--bg-input);border:1px solid var(--border);color:var(--text);border-radius:var(--rs);padding:9px 12px;font-size:13.5px;font-family:inherit;outline:none;transition:var(--tr)}
.fc:focus{border-color:var(--violet);box-shadow:0 0 0 3px rgba(124,58,237,.15)}
.fc::placeholder{color:var(--text3)}
select.fc option{background:#1a1a2e}
.tg{display:flex;align-items:center;gap:10px}
.toggle{position:relative;width:44px;height:24px}
.toggle input{opacity:0;width:0;height:0}
.ts{position:absolute;inset:0;background:rgba(255,255,255,.1);border-radius:100px;cursor:pointer;transition:var(--tr)}
.ts::before{content:'';position:absolute;left:3px;top:3px;width:18px;height:18px;border-radius:50%;background:var(--text3);transition:var(--tr)}
.toggle input:checked+.ts{background:var(--violet)}
.toggle input:checked+.ts::before{left:23px;background:#fff}

/* MODAL */
.mo{display:none;position:fixed;inset:0;z-index:1000;background:rgba(0,0,0,.7);backdrop-filter:blur(4px);align-items:center;justify-content:center}
.mo.open{display:flex;animation:fIn .2s ease}
.mb{background:#111120;border:1px solid var(--border);border-radius:16px;width:90%;max-width:560px;max-height:88vh;display:flex;flex-direction:column;box-shadow:0 20px 60px rgba(0,0,0,.6);animation:sUp .25s ease}
.mb.lg{max-width:880px}.mb.xl{max-width:1100px}
@keyframes sUp{from{transform:translateY(20px);opacity:0}to{transform:translateY(0);opacity:1}}
.mh{padding:18px 22px;border-bottom:1px solid var(--border);display:flex;align-items:center;justify-content:space-between;flex-shrink:0}
.mt{font-size:15px;font-weight:600}
.mc{background:none;border:none;color:var(--text3);cursor:pointer;font-size:20px;padding:2px;line-height:1;transition:var(--tr)}
.mc:hover{color:var(--text)}
.mbody{padding:22px;overflow-y:auto;flex:1}
.mf{padding:14px 22px;border-top:1px solid var(--border);display:flex;gap:10px;justify-content:flex-end;flex-shrink:0}

/* LOG CONSOLE */
.lc{background:#020208;border:1px solid var(--border);border-radius:var(--r);font-family:'JetBrains Mono',monospace;font-size:12px;line-height:1.7;padding:14px;overflow-y:auto;max-height:380px;min-height:180px;color:#8fffcb}
.lts{color:var(--text3)}.lok{color:var(--teal)}.lwarn{color:var(--amber)}.lerr{color:var(--red)}.linf{color:#60a5fa}

/* TOAST */
.tc{position:fixed;bottom:24px;right:24px;z-index:9999;display:flex;flex-direction:column;gap:8px}
.toast{background:#1a1a2e;border:1px solid var(--border);border-radius:var(--rs);padding:12px 18px;font-size:13px;min-width:260px;max-width:380px;display:flex;align-items:flex-start;gap:10px;box-shadow:0 4px 24px rgba(0,0,0,.4);animation:sIR .3s ease}
.toast.success{border-left:3px solid var(--green)}.toast.error{border-left:3px solid var(--red)}
.toast.info{border-left:3px solid var(--blue)}.toast.warning{border-left:3px solid var(--amber)}
@keyframes sIR{from{transform:translateX(30px);opacity:0}to{transform:translateY(0);opacity:1}}

/* SPINNER */
.sp{width:18px;height:18px;border:2px solid rgba(255,255,255,.2);border-top-color:var(--violet);border-radius:50%;animation:spin .7s linear infinite;flex-shrink:0}
@keyframes spin{to{transform:rotate(360deg)}}
.lr td{text-align:center;padding:40px;color:var(--text3)}

/* TABS */
.tabs{display:flex;gap:0;border-bottom:1px solid var(--border);margin-bottom:24px}
.tab-btn{padding:10px 18px;font-size:13px;font-weight:500;color:var(--text2);background:none;border:none;cursor:pointer;border-bottom:2px solid transparent;transition:var(--tr);font-family:inherit}
.tab-btn:hover{color:var(--text)}
.tab-btn.active{color:var(--text);border-bottom-color:var(--violet)}
.tc2{display:none}
.tc2.active{display:block;animation:fIn .2s ease}

/* TASK BAR */
.tp{background:var(--bg-card);border:1px solid var(--border);border-radius:var(--r);padding:14px 18px;margin-bottom:16px;display:flex;align-items:center;gap:14px}
#atb{display:none}.#atb.vis{display:block;margin-bottom:20px}

/* LOG FILE LIST */
.lfl{display:flex;flex-direction:column;gap:6px;margin-bottom:16px}
.lfi{display:flex;align-items:center;gap:12px;padding:10px 14px;background:var(--bg-input);border:1px solid var(--border);border-radius:var(--rs);cursor:pointer;transition:var(--tr);font-size:13px}
.lfi:hover{border-color:var(--violet);background:rgba(124,58,237,.08)}

/* IMPORT */
.ia{width:100%;min-height:180px;background:var(--bg-input);border:1px solid var(--border);color:var(--teal);border-radius:var(--rs);padding:12px;font-size:13px;font-family:'JetBrains Mono',monospace;line-height:1.6;outline:none;resize:vertical;transition:var(--tr)}
.ia:focus{border-color:var(--violet)}
.ia::placeholder{color:var(--text3);font-family:inherit}

/* DEVICE HEADER */
.dh{background:var(--bg-card);border:1px solid var(--border);border-radius:var(--r);padding:20px 24px;margin-bottom:20px;display:flex;align-items:center;justify-content:space-between;gap:16px;flex-wrap:wrap}
.dt{font-size:20px;font-weight:700;margin-bottom:4px}
.dm{font-size:12px;color:var(--text2);display:flex;gap:16px;flex-wrap:wrap}

@media(max-width:900px){.sidebar{transform:translateX(-100%)}.main{margin-left:0}.sg{grid-template-columns:repeat(2,1fr)}}
@media(max-width:600px){.sg{grid-template-columns:1fr}}
.divider{height:1px;background:var(--border);margin:20px 0}
.fw6{font-weight:600}.text-t{color:var(--teal)}.text-v{color:#a78bfa}.text-m{color:var(--text3)}.text-g{color:var(--green)}
.flex{display:flex}.aic{align-items:center}.ml-a{margin-left:auto}
.mb16{margin-bottom:16px}.mb8{margin-bottom:8px}
.trunc{overflow:hidden;text-overflow:ellipsis;white-space:nowrap;max-width:200px}
</style>
</head>
<body>
<div class="app">
<aside class="sidebar" id="sidebar">
  <div class="sb-logo">
    <div class="sb-icon">🌐</div>
    <div><div class="sb-logo-t">OOB Manager</div><div class="sb-logo-s">Network Control Panel</div></div>
  </div>
  <nav class="sb-nav">
    <div class="sb-sec">
      <div class="sb-sec-lbl">Tổng quan</div>
      <button class="nav-item active" data-page="dashboard" onclick="sPage('dashboard',this)">
        <svg class="nav-ic" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><rect x="3" y="3" width="7" height="7" rx="1"/><rect x="14" y="3" width="7" height="7" rx="1"/><rect x="3" y="14" width="7" height="7" rx="1"/><rect x="14" y="14" width="7" height="7" rx="1"/></svg>
        Dashboard
      </button>
      <button class="nav-item" data-page="devices" onclick="sPage('devices',this)">
        <svg class="nav-ic" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><rect x="2" y="6" width="20" height="12" rx="2"/><path d="M8 12h8M8 9h8M8 15h5"/></svg>
        Thiết bị OOB
        <span class="nav-badge" id="sAlBadge" style="display:none">!</span>
      </button>
    </div>
    
    {% if is_admin %}
    <div class="sb-sec">
      <div class="sb-sec-lbl">Vận hành</div>
      <button class="nav-item" data-page="verify" onclick="sPage('verify',this)">
        <svg class="nav-ic" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M9 12l2 2 4-4"/><circle cx="12" cy="12" r="9"/></svg>
        Verify & Scan
      </button>
      <button class="nav-item" data-page="logs" onclick="sPage('logs',this)">
        <svg class="nav-ic" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><polyline points="14 2 14 8 20 8"/><line x1="8" y1="13" x2="16" y2="13"/><line x1="8" y1="17" x2="16" y2="17"/></svg>
        Nhật ký Verify
      </button>
    </div>
    <div class="sb-sec">
      <div class="sb-sec-lbl">Quản lý</div>
      <button class="nav-item" data-page="import" onclick="sPage('import',this)">
        <svg class="nav-ic" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="17 8 12 3 7 8"/><line x1="12" y1="3" x2="12" y2="15"/></svg>
        Import / Export
      </button>
      <button class="nav-item" data-page="settings" onclick="sPage('settings',this)">
        <svg class="nav-ic" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><circle cx="12" cy="12" r="3"/><path d="M19.4 15a1.65 1.65 0 0 0 .33 1.82l.06.06a2 2 0 0 1-2.83 2.83l-.06-.06a1.65 1.65 0 0 0-1.82-.33 1.65 1.65 0 0 0-1 1.51V21a2 2 0 0 1-4 0v-.09A1.65 1.65 0 0 0 9 19.4a1.65 1.65 0 0 0-1.82.33l-.06.06a2 2 0 0 1-2.83-2.83l.06-.06A1.65 1.65 0 0 0 4.68 15a1.65 1.65 0 0 0-1.51-1H3a2 2 0 0 1 0-4h.09A1.65 1.65 0 0 0 4.6 9a1.65 1.65 0 0 0-.33-1.82l-.06-.06a2 2 0 0 1 2.83-2.83l.06.06A1.65 1.65 0 0 0 9 4.68a1.65 1.65 0 0 0 1-1.51V3a2 2 0 0 1 4 0v.09a1.65 1.65 0 0 0 1 1.51 1.65 1.65 0 0 0 1.82-.33l.06-.06a2 2 0 0 1 2.83 2.83l-.06.06A1.65 1.65 0 0 0 19.4 9a1.65 1.65 0 0 0 1.51 1H21a2 2 0 0 1 0 4h-.09a1.65 1.65 0 0 0-1.51 1z"/></svg>
        Cài đặt
      </button>
    </div>
    {% else %}
    <div class="sb-sec">
      <div class="sb-sec-lbl">Lịch sử & Báo cáo</div>
      <button class="nav-item" data-page="logs" onclick="sPage('logs',this)">
        <svg class="nav-ic" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><polyline points="14 2 14 8 20 8"/><line x1="8" y1="13" x2="16" y2="13"/><line x1="8" y1="17" x2="16" y2="17"/></svg>
        Nhật ký Verify
      </button>
      <button class="nav-item" data-page="import" onclick="sPage('import',this)">
        <svg class="nav-ic" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="17 8 12 3 7 8"/><line x1="12" y1="3" x2="12" y2="15"/></svg>
        Xuất Excel
      </button>
    </div>
    {% endif %}
  </nav>
  <div class="sb-foot">
    <div class="flex aic" style="gap:6px"><div class="d-dot" id="dDot"></div><span id="dTxt">Đang kiểm tra...</span></div>
    <div style="margin-top:6px;font-size:10px" id="clk"></div>
  </div>
</aside>

<div class="main">
  <div class="topbar">
    <div class="tb-title" id="tbTitle">Dashboard</div>
    <div class="sw">
      <svg class="si" width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><circle cx="11" cy="11" r="8"/><path d="m21 21-4.35-4.35"/></svg>
      <input type="text" id="gSearch" placeholder="Tìm IP, hostname, description..." onkeydown="if(event.key==='Enter')doSearch()">
    </div>
    <button class="btn btn-p" onclick="doSearch()" style="padding:8px 14px; margin-right:auto">
      <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5"><circle cx="11" cy="11" r="8"/><path d="m21 21-4.35-4.35"/></svg>Tìm
    </button>
    
    {% if is_admin %}
    <button class="btn btn-g btn-sm" onclick="oModal('addDev')">+ Thêm OOB</button>
    <button class="btn btn-outline-info btn-sm me-2" onclick="sPage('settings')"><i class="bi bi-gear"></i> Cài đặt</button>
    <button class="btn btn-a btn-sm" onclick="oModal('changePassMod')" style="margin-left:8px">🔑 Đổi Pass Web</button>
    <button class="btn btn-d btn-sm" onclick="window.location.href='/logout'" style="margin-left:8px">Đăng xuất</button>
    {% else %}
    <button class="btn btn-p btn-sm" onclick="window.location.href='/login'" style="margin-left:8px">Đăng nhập Quản trị</button>
    {% endif %}
  </div>
  
  <div class="content">
    <div id="atb"></div>

    <!-- DASHBOARD -->
    <div class="page active" id="page-dashboard">
      <div style="display:flex;gap:24px;margin-bottom:24px;flex-wrap:wrap">
        <div class="sg" style="flex:1;margin-bottom:0;min-width:400px">
          <div class="sc c1"><div class="sv" id="s-total">-</div><div class="sl">Tổng thiết bị</div><div class="si2">🌐</div></div>
          <div class="sc c2"><div class="sv" id="s-online">-</div><div class="sl">Đang online</div><div class="si2">📶</div></div>
          <div class="sc c3"><div class="sv" id="s-baseline">-</div><div class="sl">Có baseline</div><div class="si2">📋</div></div>
          <div class="sc c4"><div class="sv" id="s-alarms">-</div><div class="sl">Cảnh báo</div><div class="si2">⚠️</div></div>
        </div>
        <div style="display:flex;gap:16px;flex-wrap:wrap">
          <div class="sc" style="width:200px;text-align:center;padding:15px;display:flex;flex-direction:column;align-items:center">
            <div class="st mb8"><span class="dot" style="background:var(--teal)"></span>Ping Status</div>
            <svg viewBox="0 0 36 36" style="width:100px;height:100px;margin-top:10px">
              <path class="circle-bg" d="M18 2.0845 a 15.9155 15.9155 0 0 1 0 31.831 a 15.9155 15.9155 0 0 1 0 -31.831" fill="none" stroke="rgba(255,255,255,0.1)" stroke-width="3.5" />
              <path id="svg-ping" d="M18 2.0845 a 15.9155 15.9155 0 0 1 0 31.831 a 15.9155 15.9155 0 0 1 0 -31.831" fill="none" stroke="var(--teal)" stroke-width="3.5" stroke-dasharray="0, 100" style="transition:stroke-dasharray 1s ease"/>
              <text id="svg-ping-txt" x="18" y="21.5" fill="var(--text)" font-size="9" text-anchor="middle" font-weight="600">0%</text>
            </svg>
          </div>
          <div class="sc" style="width:200px;text-align:center;padding:15px;display:flex;flex-direction:column;align-items:center">
            <div class="st mb8"><span class="dot" style="background:var(--amber)"></span>Baseline</div>
            <svg viewBox="0 0 36 36" style="width:100px;height:100px;margin-top:10px">
              <path class="circle-bg" d="M18 2.0845 a 15.9155 15.9155 0 0 1 0 31.831 a 15.9155 15.9155 0 0 1 0 -31.831" fill="none" stroke="rgba(255,255,255,0.1)" stroke-width="3.5" />
              <path id="svg-baseline" d="M18 2.0845 a 15.9155 15.9155 0 0 1 0 31.831 a 15.9155 15.9155 0 0 1 0 -31.831" fill="none" stroke="var(--amber)" stroke-width="3.5" stroke-dasharray="0, 100" style="transition:stroke-dasharray 1s ease"/>
              <text id="svg-baseline-txt" x="18" y="21.5" fill="var(--text)" font-size="9" text-anchor="middle" font-weight="600">0%</text>
            </svg>
          </div>
        </div>
      </div>
      <div class="sh">
        <div class="st"><span class="dot"></span>Thiết bị OOB</div>
        <div class="bg2">
          {% if is_admin %}
          <button class="btn btn-t btn-sm" onclick="runAction('scan',null)">🔍 Scan All</button>
          <button class="btn btn-a btn-sm" onclick="runAction('verify',null)">⚡ Verify All</button>
          <button class="btn btn-pk btn-sm" onclick="runAction('push',null)">🚀 Push All</button>
          {% endif %}
          <button class="btn btn-g btn-sm" onclick="loadDash()">↻ Làm mới</button>
        </div>
      </div>
      <div class="tw">
        <table id="dashTable">
          <thead><tr><th>Alias</th><th>IP</th><th>Hostname</th><th>Ping</th><th>Menu</th><th style="text-align:center">Lines</th><th>Verify</th><th>Cập nhật</th><th style="text-align:right">Hành động</th></tr></thead>
          <tbody id="dashBody"><tr class="lr"><td colspan="9"><div class="sp" style="margin:0 auto"></div></td></tr></tbody>
        </table>
      </div>
    </div>

    <!-- DEVICES -->
    <div class="page" id="page-devices">
      <div class="dh" id="devHdr" style="display:none">
        <div>
          <div class="dt" id="devTitle">-</div>
          <div class="dm">
            <span>📡 IP: <strong id="devIP">-</strong></span>
            <span>🖧 Menu: <strong id="devMenu">-</strong></span>
            <span>📦 Lines: <strong id="devLines">-</strong></span>
          </div>
        </div>
        <div class="bg2">
          {% if is_admin %}
          <button class="btn btn-t btn-sm" onclick="runAction('scan',curIP)">🔍 Scan</button>
          <button class="btn btn-a btn-sm" onclick="runAction('verify',curIP)">⚡ Verify</button>
          <button class="btn btn-pk btn-sm" onclick="runAction('push',curIP)">🚀 Push</button>
          {% endif %}
          <button class="btn btn-g btn-sm" onclick="sPage('dashboard')">← Quay lại</button>
        </div>
      </div>
      <div class="tw">
        <table>
          <thead><tr><th>Option Key</th><th>Description</th><th>Target IP</th><th>Port</th><th>Protocol</th><th>Verify</th><th>Hostname Thực tế</th></tr></thead>
          <tbody id="devOptsBody"><tr class="lr"><td colspan="7"><div class="sp" style="margin:0 auto"></div></td></tr></tbody>
        </table>
      </div>
    </div>

    {% if is_admin %}
    <!-- VERIFY & SCAN (Chi Admin) -->
    <div class="page" id="page-verify">
      <div class="sh mb16"><div class="st"><span class="dot"></span>Vận hành Tức thì</div></div>
      <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:16px;margin-bottom:24px">
        <div class="sc c1" style="cursor:pointer" onclick="runAction('scan',null)"><div style="font-size:28px;margin-bottom:10px">🔍</div><div class="sl">SCAN CONFIG</div><div style="font-size:12px;color:var(--text2);margin-top:6px">Thu thập cấu hình menu từ tất cả OOB</div></div>
        <div class="sc c3" style="cursor:pointer" onclick="runAction('verify',null)"><div style="font-size:28px;margin-bottom:10px">⚡</div><div class="sl">DEEP VERIFY</div><div style="font-size:12px;color:var(--text2);margin-top:6px">Kiểm tra vật lý PIVOT tất cả line console</div></div>
        <div class="sc c4" style="cursor:pointer" onclick="runAction('push',null)"><div style="font-size:28px;margin-bottom:10px">🚀</div><div class="sl">PUSH CONFIG</div><div style="font-size:12px;color:var(--text2);margin-top:6px">Tự động sửa Description sai lệch</div></div>
      </div>
      <div class="sh"><div class="st"><span class="dot"></span>Chạy cho thiết bị cụ thể</div></div>
      <div style="background:var(--bg-card);border:1px solid var(--border);border-radius:var(--r);padding:20px;margin-bottom:24px">
        <div class="fg"><label class="fl">IP hoặc Alias (để trống = Tất cả)</label><input type="text" id="specIP" class="fc" placeholder="VD: 192.168.1.1"></div>
        <div class="bg2">
          <button class="btn btn-t" onclick="runAction('scan',document.getElementById('specIP').value.trim()||null)">🔍 Scan Device</button>
          <button class="btn btn-a" onclick="runAction('verify',document.getElementById('specIP').value.trim()||null)">⚡ Verify Device</button>
          <button class="btn btn-pk" onclick="runAction('push',document.getElementById('specIP').value.trim()||null)">🚀 Push Device</button>
        </div>
      </div>
      <div class="sh"><div class="st"><span class="dot"></span>Live Console</div><button class="btn btn-g btn-sm" onclick="document.getElementById('liveCon').innerHTML='<span class=text-m>Console da xoa.</span>'">Xóa</button></div>
      <div class="lc" id="liveCon" style="min-height:260px;max-height:500px"><span class="text-m">Chờ lệnh...</span></div>
    </div>
    {% endif %}

    <!-- LOGS -->
    <div class="page" id="page-logs">
      <div class="tabs mb16">
        <button class="tab-btn active" id="btn-vlogs" onclick="sTabLogs('vlogs',this)">📋 Nhật ký Verify</button>
        <button class="tab-btn" id="btn-plogs" onclick="sTabLogs('plogs',this)">🚀 Lịch sử Push & Revert</button>
        <button class="btn btn-g btn-sm ml-a" onclick="loadLogs()">↻ Làm mới</button>
      </div>
      <div style="display:grid;grid-template-columns:300px 1fr;gap:16px">
        <div><div class="lfl" id="logList"><div class="text-m" style="padding:12px;font-size:13px">Đang tải...</div></div></div>
        <div>
          <div class="lc" id="logView" style="max-height:70vh;min-height:300px;font-size:11.5px;white-space:pre-wrap;word-break:break-all"><span class="text-m">← Chọn file log bên trái để xem.</span></div>
          {% if is_admin %}
          <div id="revertAction" style="display:none;margin-top:10px;text-align:right">
            <button class="btn btn-d" id="btnRevert" onclick="doRevert()">↩ Revert Changes</button>
          </div>
          {% endif %}
        </div>
      </div>
    </div>

    <!-- IMPORT / EXPORT -->
    <div class="page" id="page-import">
      <div style="display:grid;grid-template-columns:{% if is_admin %}1fr 1fr{% else %}1fr{% endif %};gap:20px">
        {% if is_admin %}
        <div style="background:var(--bg-card);border:1px solid var(--border);border-radius:var(--r);padding:22px">
          <div class="st mb16"><span class="dot"></span>Import Danh sách IP</div>
          <p style="font-size:13px;color:var(--text2);margin-bottom:14px">Mỗi dòng 1 thiết bị. Format: <code style="color:var(--teal)">IP [alias]</code></p>
          <textarea id="importTxt" class="ia" placeholder="192.168.1.1 OOB-HCM-01&#10;192.168.1.2 OOB-HCM-02"></textarea>
          <div style="margin-top:14px"><button class="btn btn-p" onclick="doImport()">⬆ Import</button></div>
          <div id="importRes" style="margin-top:12px;font-size:13px;display:none"></div>
        </div>
        {% endif %}
        <div style="background:var(--bg-card);border:1px solid var(--border);border-radius:var(--r);padding:22px">
          <div class="st mb16"><span class="dot"></span>Xuất Báo cáo Excel</div>
          <p style="font-size:13px;color:var(--text2);margin-bottom:14px">Xuất toàn bộ dữ liệu baseline và kết quả verify ra .xlsx</p>
          <div style="background:rgba(6,214,160,.06);border:1px solid rgba(6,214,160,.2);border-radius:var(--rs);padding:14px;margin-bottom:20px;font-size:12.5px;color:var(--text2)">
            📊 Bao gồm: Tất cả OOB + trạng thái ping · Danh sách option baseline · Kết quả verify
          </div>
          <a href="/api/export/excel" class="btn btn-t" download>⬇ Tải về Excel</a>
        </div>
      </div>
    </div>

    {% if is_admin %}
    <!-- SETTINGS -->
    <div class="page" id="page-settings">
      <div class="tabs">
        <button class="tab-btn active" onclick="sTab('tab-conn',this)">🔐 Kết nối</button>
        <button class="tab-btn" onclick="sTab('tab-multi',this)">👥 Multi-Account</button>
        <button class="tab-btn" onclick="sTab('tab-sched',this)">⏱️ Lịch chạy</button>
        <button class="tab-btn" onclick="sTab('tab-files',this)">📂 Files</button>
      </div>
      <!-- Ket noi -->
      <div class="tc2 active" id="tab-conn">
        <div style="max-width:600px">
          <div class="fg"><label class="fl">Username chính</label><input type="text" id="cu" class="fc" placeholder="admin"></div>
          <div class="fg"><label class="fl">Password chính</label><input type="password" id="cp" class="fc"></div>
          <div class="fg"><label class="fl">Enable Password</label><input type="password" id="ce" class="fc"></div>
          <div class="fg"><label class="fl">Vertiv Connect Password</label><input type="password" id="cv" class="fc"></div>
          <div style="display:grid;grid-template-columns:1fr 1fr;gap:16px">
            <div class="fg"><label class="fl">SSH Port</label><input type="number" id="csp" class="fc" placeholder="22"></div>
            <div class="fg"><label class="fl">Telnet Port</label><input type="number" id="ctp" class="fc" placeholder="23"></div>
          </div>
          <div class="fg"><label class="fl">Tên menu ép dùng (trống = tự dò)</label><input type="text" id="cmno" class="fc" placeholder="OOB_MENU"></div>
          <div class="fg"><div class="tg"><label class="toggle"><input type="checkbox" id="cav"><span class="ts"></span></label><span style="font-size:13.5px">Bật Tự động Verify ngầm</span></div></div>
          <button class="btn btn-p" onclick="saveCfg()">💾 Lưu cài đặt</button>
        </div>
      </div>
      <!-- Multi Account -->
      <div class="tc2" id="tab-multi">
        <div style="max-width:700px">
          <p style="font-size:13px;color:var(--text2);margin-bottom:18px">Tool sẽ thử lần lượt khi tài khoản chính thất bại.</p>
          <div id="credList" style="margin-bottom:18px"></div>
          <button class="btn btn-g" onclick="oModal('addCred')">+ Thêm tài khoản phụ</button>
        </div>
      </div>
      <!-- Lich chay -->
      <div class="tc2" id="tab-sched">
        <div style="max-width:600px">
          <div style="background:var(--bg-card);border:1px solid var(--border);border-radius:var(--r);padding:20px;margin-bottom:20px">
            <div class="st mb16">🔍 Lịch Thu thập Config (Scan)</div>
            <div class="fg"><label class="fl">Chế độ</label><select id="csm" class="fc" onchange="togSF('scan')"><option value="interval">Lặp lại theo chu kỳ</option><option value="daily">Hàng ngày</option><option value="weekly">Hàng tuần</option></select></div>
            <div class="fg"><label class="fl">Chu kỳ scan (giây)</label><input type="number" id="ci" class="fc" placeholder="30"></div>
            <div id="s_tf" style="display:none">
              <div style="display:grid;grid-template-columns:1fr 1fr;gap:16px">
                <div class="fg"><label class="fl">Giờ chạy (HH:MM)</label><input type="text" id="cst" class="fc" placeholder="01:00"></div>
                <div class="fg" id="s_wf" style="display:none"><label class="fl">Thứ (mon/tue.../sun)</label><input type="text" id="csw" class="fc" placeholder="mon"></div>
              </div>
            </div>
          </div>
          <div style="background:var(--bg-card);border:1px solid var(--border);border-radius:var(--r);padding:20px;margin-bottom:20px">
            <div class="st mb16">⚡ Lịch Deep Verify</div>
            <div class="fg"><label class="fl">Chế độ</label><select id="cvm" class="fc" onchange="togSF('verify')"><option value="interval">Lặp lại theo chu kỳ</option><option value="daily">Hàng ngày</option><option value="weekly">Hàng tuần</option></select></div>
            <div class="fg"><label class="fl">Chu kỳ verify (giây)</label><input type="number" id="cvi" class="fc" placeholder="3600"></div>
            <div id="v_tf" style="display:none">
              <div style="display:grid;grid-template-columns:1fr 1fr;gap:16px">
                <div class="fg"><label class="fl">Giờ chạy (HH:MM)</label><input type="text" id="cvt" class="fc" placeholder="01:00"></div>
                <div class="fg" id="v_wf" style="display:none"><label class="fl">Thứ</label><input type="text" id="cvw" class="fc" placeholder="mon"></div>
              </div>
            </div>
            <div style="display:grid;grid-template-columns:1fr 1fr;gap:16px">
              <div class="fg"><label class="fl">Chờ sau connect (s)</label><input type="number" id="cvwac" class="fc" step="0.5" placeholder="1.5"></div>
              <div class="fg"><label class="fl">Timeout Verify (s)</label><input type="number" id="cmvd" class="fc" placeholder="300"></div>
            </div>
          </div>
          <button class="btn btn-p" onclick="saveSched()">💾 Lưu lịch</button>
        </div>
      </div>
      <!-- Files -->
      <div class="tc2" id="tab-files">
        <div style="max-width:600px">
          <div class="fg"><label class="fl">File danh sách IP</label><input type="text" id="cil" class="fc" placeholder="oob_ips.txt"></div>
          <div class="fg"><label class="fl">File Baseline DB</label><input type="text" id="cbd" class="fc" placeholder="baseline.db"></div>
          <div class="fg"><label class="fl">File Snapshot DB</label><input type="text" id="csd" class="fc" placeholder="snapshot.db"></div>
          <button class="btn btn-p" onclick="saveFiles()">💾 Lưu</button>
        </div>
      </div>
    </div>
    {% endif %}

  </div>
</div>

{% if is_admin %}
<!-- MODALS ADMIN CHỈ CÓ NẾU ĐĂNG NHẬP -->
<div class="mo" id="addDev">
  <div class="mb">
    <div class="mh"><div class="mt">➕ Thêm thiết bị OOB</div><button class="mc" onclick="cModal('addDev')">×</button></div>
    <div class="mbody">
      <div class="fg"><label class="fl">IP Address *</label><input type="text" id="nip" class="fc" placeholder="192.168.1.100"></div>
      <div class="fg"><label class="fl">Alias (Tên gọi)</label><input type="text" id="nal" class="fc" placeholder="OOB-HCM-01"></div>
    </div>
    <div class="mf"><button class="btn btn-g" onclick="cModal('addDev')">Hủy</button><button class="btn btn-p" onclick="addDevice()">Thêm mới</button></div>
  </div>
</div>

<div class="mo" id="addCred">
  <div class="mb">
    <div class="mh"><div class="mt">👤 Thêm tài khoản phụ</div><button class="mc" onclick="cModal('addCred')">×</button></div>
    <div class="mbody">
      <div class="fg"><label class="fl">Username *</label><input type="text" id="cru" class="fc" placeholder="admin"></div>
      <div class="fg"><label class="fl">Password</label><input type="password" id="crp" class="fc"></div>
      <div class="fg"><label class="fl">Enable Password</label><input type="password" id="cre" class="fc"></div>
    </div>
    <div class="mf"><button class="btn btn-g" onclick="cModal('addCred')">Hủy</button><button class="btn btn-p" onclick="addCred()">Thêm</button></div>
  </div>
</div>

<!-- Modal Đổi Mật Khẩu Web -->
<div class="mo" id="changePassMod">
  <div class="mb">
    <div class="mh"><div class="mt">🔑 Đổi Mật Khẩu Đăng Nhập Web</div><button class="mc" onclick="cModal('changePassMod')">×</button></div>
    <div class="mbody">
      <div class="fg"><label class="fl">Mật khẩu cũ *</label><input type="password" id="cp_old" class="fc"></div>
      <div class="fg"><label class="fl">Mật khẩu mới *</label><input type="password" id="cp_new" class="fc"></div>
      <div class="fg"><label class="fl">Xác nhận mật khẩu mới *</label><input type="password" id="cp_cfm" class="fc"></div>
    </div>
    <div class="mf"><button class="btn btn-g" onclick="cModal('changePassMod')">Hủy</button><button class="btn btn-p" onclick="doChangePass()">Lưu Thay Đổi</button></div>
  </div>
</div>
{% endif %}

<!-- Modal Kết quả Tìm kiếm (Dùng chung cho Guest/Admin) -->
<div class="mo" id="searchMod">
  <div class="mb xl">
    <div class="mh"><div class="mt">🔍 Kết quả: "<span id="skw"></span>"</div><button class="mc" onclick="cModal('searchMod')">×</button></div>
    <div class="mbody" style="padding:0;overflow-x:auto">
      <table>
        <thead><tr><th style="padding-left:18px">OOB</th><th>Port</th><th>Description</th><th>Hostname Thực tế</th><th>Kết nối</th><th>Verify</th><th style="padding-right:18px">Đi tới</th></tr></thead>
        <tbody id="sBdy"></tbody>
      </table>
    </div>
  </div>
</div>

<div class="tc" id="toastCnt"></div>

<script>
const isAdmin = {{ 'true' if is_admin else 'false' }};
let curPage='dashboard', curIP=null, sse=null, actTasks={};

function initSSE(){
  if(sse)sse.close();
  sse=new EventSource('/api/events');
  sse.onmessage=e=>{
    try{
      const d=JSON.parse(e.data);
      if(d.type==='log')logMsg(d);
      if(d.type==='task_done')taskDone(d);
    }catch{}
  };
  sse.onerror=()=>setTimeout(initSSE,5000);
}

function logMsg(d){
  const con=document.getElementById('liveCon');
  if(!con) return;
  const msg=esc(d.msg||'');
  let cls='linf';
  if(/OK|thanh cong|khop/i.test(msg))cls='lok';
  else if(/LOI|that bai|error/i.test(msg))cls='lerr';
  else if(/CANH BAO|warn/i.test(msg))cls='lwarn';
  con.innerHTML+=`<div><span class="lts">[${d.ts||''}]</span> <span class="${cls}">${msg}</span></div>`;
  con.scrollTop=con.scrollHeight;
}

function taskDone(d){
  const tid=d.task;
  if(actTasks[tid]){
    delete actTasks[tid]; renderTasks();
    toast(tid.split('_')[0].toUpperCase()+' hoàn thành!','success');
    if(curPage==='dashboard')loadDash();
    if(curPage==='devices'&&curIP)loadDevOpts(curIP);
  }
}

function addTask(tid,lbl){actTasks[tid]={label:lbl};renderTasks();}
function renderTasks(){
  const bar=document.getElementById('atb');
  const ts=Object.entries(actTasks);
  if(!ts.length){bar.className='';bar.innerHTML='';return;}
  bar.className='vis';
  bar.innerHTML=ts.map(([id,t])=>`<div class="tp"><div class="sp"></div><div style="flex:1"><div style="font-size:13px;font-weight:500">${esc(t.label)}</div><div style="font-size:11px;color:var(--text3)">Đang chạy...</div></div><span class="badge bv bpulse">RUNNING</span></div>`).join('');
}

const ptitles={dashboard:'Dashboard',devices:'Chi tiết thiết bị',verify:'Verify & Scan',logs:'Nhật ký Verify',import:'Import / Export',settings:'Cài đặt'};
function sPage(page,btn){
  document.querySelectorAll('.page').forEach(p=>p.classList.remove('active'));
  document.querySelectorAll('.nav-item').forEach(n=>n.classList.remove('active'));
  const el=document.getElementById('page-'+page);
  if(el)el.classList.add('active');
  if(btn)btn.classList.add('active');
  else{const n=document.querySelector('[data-page="'+page+'"]');if(n)n.classList.add('active');}
  curPage=page;
  document.getElementById('tbTitle').textContent=ptitles[page]||page;
  if(page==='dashboard')loadDash();
  if(page==='logs')loadLogs();
  if(isAdmin && page==='settings')loadSettings();
}

function sTab(tabId,btn){
  document.querySelectorAll('.tc2').forEach(t=>t.classList.remove('active'));
  document.querySelectorAll('.tab-btn').forEach(b=>b.classList.remove('active'));
  document.getElementById(tabId).classList.add('active');
  btn.classList.add('active');
  if(tabId==='tab-multi')loadCreds();
}

async function loadDash(){
  const[stats,devs]=await Promise.all([fetch('/api/stats').then(r=>r.json()).catch(()=>({})),fetch('/api/devices').then(r=>r.json()).catch(()=>[])]);
  document.getElementById('s-total').textContent=stats.total??0;
  document.getElementById('s-online').textContent=stats.online??0;
  document.getElementById('s-baseline').textContent=stats.has_baseline??0;
  document.getElementById('s-alarms').textContent=stats.alarms??0;
  
  const pctPing=stats.total>0?Math.round(((stats.online||0)/stats.total)*100):0;
  document.getElementById('svg-ping').style.strokeDasharray=pctPing+', 100';
  document.getElementById('svg-ping-txt').textContent=pctPing+'%';
  const pctBase=stats.total>0?Math.round(((stats.has_baseline||0)/stats.total)*100):0;
  document.getElementById('svg-baseline').style.strokeDasharray=pctBase+', 100';
  document.getElementById('svg-baseline-txt').textContent=pctBase+'%';
  
  const ab=document.getElementById('sAlBadge');
  if(stats.alarms>0){ab.style.display='';ab.textContent=stats.alarms;}else ab.style.display='none';
  const tbody=document.getElementById('dashBody');
  if(!devs.length){tbody.innerHTML='<tr><td colspan="9" style="text-align:center;padding:50px;color:var(--text3)">Chưa có thiết bị. Cần Đăng nhập Quản trị để thêm mới.</td></tr>';return;}
  
  tbody.innerHTML=devs.map(d=>{
    const pb=d.ping===true?'<span class="badge bg">● Online</span>':d.ping===false?'<span class="badge br">● Offline</span>':'<span class="badge bm">- Chưa</span>';
    const mb=d.menu_state==='ok'?'<span class="badge bg">OK</span>':d.menu_state==='conn_failed'?'<span class="badge br">Lỗi</span>':d.menu_state==='no_menu'?'<span class="badge ba">No Menu</span>':'<span class="badge bm">'+(d.menu_state||'-')+'</span>';
    const ab2=d.alarm_count>0?'<span class="badge bp">'+(d.alarm_count)+' ⚠️</span>':d.ok_count>0?'<span class="badge bt">'+d.ok_count+' ✓</span>':'<span class="badge bm">-</span>';
    const upd=(d.updated_at||d.checked_at||'-').replace('T',' ').slice(0,16);
    
    let actHtml = `<button class="btn btn-g btn-sm btn-ic" onclick="openDev('${esc(d.ip)}','${esc(d.alias)}')" title="Chi tiết Line">👁</button>`;
    if(isAdmin) {
        actHtml += `
          <button class="btn btn-t btn-sm btn-ic" onclick="runAction('scan','${esc(d.ip)}')" title="Scan">🔍</button>
          <button class="btn btn-a btn-sm btn-ic" onclick="runAction('verify','${esc(d.ip)}')" title="Verify">⚡</button>
          <button class="btn btn-pk btn-sm btn-ic" onclick="runAction('push','${esc(d.ip)}')" title="Push">🚀</button>
          <button class="btn btn-d btn-sm btn-ic" onclick="delDev('${esc(d.ip)}')" title="Xóa">🗑</button>
        `;
    }

    return`<tr>
      <td><span class="fw6">${esc(d.alias)}</span></td>
      <td><span class="mono">${esc(d.ip)}</span></td>
      <td><span class="text-m" style="font-size:12px">${esc(d.device_name||'-')}</span></td>
      <td>${pb}</td><td>${mb}</td>
      <td style="text-align:center"><span class="badge bv">${d.opt_count}</span></td>
      <td>${ab2}</td>
      <td style="font-size:11px;color:var(--text3)">${esc(upd)}</td>
      <td style="text-align:right">
        <div class="bg2" style="justify-content:flex-end">
          ${actHtml}
        </div>
      </td></tr>`;
  }).join('');
}

function openDev(ip,alias){
  curIP=ip; sPage('devices');
  document.getElementById('devHdr').style.display='';
  document.getElementById('devTitle').textContent=alias||ip;
  document.getElementById('devIP').textContent=ip;
  loadDevOpts(ip);
}

async function loadDevOpts(ip){
  document.getElementById('devOptsBody').innerHTML='<tr class="lr"><td colspan="7"><div class="sp" style="margin:0 auto"></div></td></tr>';
  const d=await fetch('/api/device/'+encodeURIComponent(ip)+'/options').then(r=>r.json()).catch(()=>null);
  if(!d){document.getElementById('devOptsBody').innerHTML='<tr><td colspan="7" style="text-align:center;color:var(--text3)">Lỗi tải dữ liệu</td></tr>';return;}
  document.getElementById('devMenu').textContent=d.menu_name||'-';
  document.getElementById('devLines').textContent=d.options.length;
  if(!d.options.length){document.getElementById('devOptsBody').innerHTML='<tr><td colspan="7" style="text-align:center;padding:40px;color:var(--text3)">Chưa có baseline. Vui lòng liên hệ Admin quét cấu hình.</td></tr>';return;}
  document.getElementById('devOptsBody').innerHTML=d.options.map(o=>{
    const pc=o.protocol==='ssh'?'bt':o.protocol==='serial'?'bv':'ba';
    let vb='<span class="badge bm">-</span>';
    if(o.verify_status==='OK')vb='<span class="badge bg">✓ OK</span>';
    else if(o.verify_status==='CANH BAO')vb='<span class="badge bp bpulse">⚠ Cảnh báo</span>';
    else if(o.verify_status==='TIMEOUT')vb='<span class="badge ba">⏱ Timeout</span>';
    else if(o.verify_status==='KHONG PIVOT')vb='<span class="badge br">↩ No Pivot</span>';
    else if(o.verify_status==='YEU CAU DANG NHAP')vb='<span class="badge ba">🔑 Auth</span>';
    const ah=o.act_host?'<span class="text-t fw6">'+esc(o.act_host)+'</span>':'<span class="text-m">-</span>';
    return`<tr>
      <td><kbd style="background:rgba(124,58,237,.2);color:#c4b5fd;border-radius:4px;padding:2px 8px;font-family:'JetBrains Mono',monospace;font-size:12px">${esc(o.key)}</kbd></td>
      <td>${esc(o.description)}</td>
      <td><span class="mono text-t">${esc(o.ip)}</span></td>
      <td><span class="mono">${o.port}</span></td>
      <td><span class="badge ${pc}">${esc(o.protocol.toUpperCase())}</span></td>
      <td>${vb}</td><td>${ah}</td></tr>`;
  }).join('');
}

async function runAction(action,ip){
  if(!isAdmin) { toast('Bạn cần Đăng nhập Quản trị để thực hiện lệnh này!','error'); return; }
  const lbl=action.toUpperCase()+' '+(ip||'Tất cả');
  if(!confirm('Xác nhận chạy lệnh '+lbl+'?'))return;
  if(curPage!=='verify')sPage('verify');
  const con=document.getElementById('liveCon');
  con.innerHTML+='<div><span class="lts">['+nw()+']</span> <span class="linf">▶ Bắt đầu '+esc(lbl)+'...</span></div>';
  con.scrollTop=con.scrollHeight;
  const res=await fetch('/api/action',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({action,ip})}).then(r=>r.json()).catch(()=>null);
  if(res&&res.task_id){addTask(res.task_id,lbl);toast(res.msg||'Đã đưa vào hàng đợi!','info');}
  else toast('Lỗi gửi lệnh! Bạn có thể đã hết phiên đăng nhập.','error');
}

async function addDevice(){
  if(!isAdmin) return;
  const ip=document.getElementById('nip').value.trim(),alias=document.getElementById('nal').value.trim();
  if(!ip){toast('Vui lòng nhập IP!','error');return;}
  await fetch('/api/device',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({ip,alias})});
  cModal('addDev');toast('Đã thêm '+ip+'!','success');loadDash();
}

async function delDev(ip){
  if(!isAdmin) return;
  if(!confirm('Xóa OOB '+ip+'?'))return;
  await fetch('/api/device',{method:'DELETE',headers:{'Content-Type':'application/json'},body:JSON.stringify({ip})});
  toast('Đã xóa '+ip+'!','success');loadDash();
}

async function doChangePass(){
  if(!isAdmin) return;
  const op=g('cp_old').value, np=g('cp_new').value, cp=g('cp_cfm').value;
  if(!op||!np||!cp){toast('Vui lòng điền đủ thông tin!','error');return;}
  if(np!==cp){toast('Mật khẩu mới không khớp!','error');return;}
  
  const r=await fetch('/api/change-password',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({old_password:op,new_password:np})}).then(x=>x.json()).catch(()=>null);
  if(r&&r.status==='ok'){
    toast(r.msg,'success');
    cModal('changePassMod');
    g('cp_old').value='';g('cp_new').value='';g('cp_cfm').value='';
  }else toast((r&&r.msg)||'Lỗi đổi mật khẩu!','error');
}

async function doSearch(){
  const q=document.getElementById('gSearch').value.trim();
  if(!q)return;
  document.getElementById('skw').textContent=q;
  document.getElementById('sBdy').innerHTML='<tr class="lr"><td colspan="7"><div class="sp" style="margin:0 auto"></div></td></tr>';
  oModal('searchMod');
  const data=await fetch('/api/search?q='+encodeURIComponent(q)).then(r=>r.json()).catch(()=>[]);
  if(!data.length){document.getElementById('sBdy').innerHTML='<tr><td colspan="7" style="text-align:center;padding:40px;color:var(--text3)">Không tìm thấy kết quả cho "'+esc(q)+'"</td></tr>';return;}
  const vbadge=s=>{if(!s)return'';if(s==='OK')return'<span class="badge bg">✓</span>';if(s==='CANH BAO')return'<span class="badge bp">⚠️</span>';return'<span class="badge bm">'+s+'</span>';};
  document.getElementById('sBdy').innerHTML=data.map(item=>`
    <tr>
      <td style="padding-left:18px"><strong class="text-t">${esc(item.oob_alias)}</strong><br><small class="text-m mono">${esc(item.oob_ip)}</small></td>
      <td><kbd style="background:rgba(124,58,237,.2);color:#c4b5fd;padding:2px 8px;border-radius:4px;font-family:monospace">${esc(item.opt_key)}</kbd></td>
      <td>${esc(item.desc)}</td>
      <td>${item.act_host?'<span class="text-g fw6">'+esc(item.act_host)+'</span>':'<span class="text-m">-</span>'}</td>
      <td><span class="mono text-m" style="font-size:11px">${esc(item.protocol)}://${esc(item.target_ip)}:${item.target_port}</span></td>
      <td>${vbadge(item.verify_status)}</td>
      <td style="padding-right:18px"><button class="btn btn-g btn-sm" onclick="cModal('searchMod');openDev('${esc(item.oob_ip)}','${esc(item.oob_alias)}')">Đi tới →</button></td>
    </tr>`).join('');
}

let curLogTab='vlogs', curLogFile='';
function sTabLogs(tab,btn){
  document.querySelectorAll('#page-logs .tab-btn').forEach(b=>b.classList.remove('active'));
  btn.classList.add('active');
  curLogTab=tab;
  loadLogs();
}

async function loadLogs(){
  const isP = curLogTab === 'plogs';
  const api = isP ? '/api/push-logs' : '/api/logs';
  const files=await fetch(api).then(r=>r.json()).catch(()=>[]);
  const el=document.getElementById('logList');
  if(!files.length){el.innerHTML='<div style="padding:16px;color:var(--text3);font-size:13px">Chưa có file log.</div>';return;}
  el.innerHTML=files.map(f=>`<div class="lfi" onclick="loadLogCnt('${encodeURIComponent(f.name)}')"><span style="font-size:16px">${isP?'🚀':'📄'}</span><div style="flex:1;overflow:hidden"><div class="mono" style="overflow:hidden;text-overflow:ellipsis;white-space:nowrap;font-size:11px">${esc(f.name)}</div><div style="font-size:10px;color:var(--text3)">${f.mtime}</div></div></div>`).join('');
  const rAction = document.getElementById('revertAction');
  if(rAction) rAction.style.display='none';
  document.getElementById('logView').innerHTML='<span class="text-m">← Chọn file log bên trái để xem.</span>';
}

async function loadLogCnt(fn){
  curLogFile = fn;
  document.getElementById('logView').textContent='Đang tải...';
  const api = (curLogTab === 'plogs' ? '/api/push-logs/' : '/api/logs/') + fn;
  const d=await fetch(api).then(r=>r.json()).catch(()=>null);
  let txt = d&&d.content?d.content:'Lỗi tải file.';
  txt = esc(txt);
  txt = txt.replace(/OK|thành công/gi, '<span class="lok">$&</span>')
           .replace(/CANH BAO|KHONG PIVOT|TIMEOUT|YEU CAU DANG NHAP/g, '<span class="lwarn">$&</span>')
           .replace(/LOI|LỖI/g, '<span class="lerr">$&</span>')
           .replace(/REVERT CMD/g, '<span class="bpulse" style="color:var(--pink)">$&</span>');
  document.getElementById('logView').innerHTML=txt;
  
  const rAction = document.getElementById('revertAction');
  if(isAdmin && rAction) {
    if(curLogTab === 'plogs' && d && d.content && txt.includes('REVERT CMD')) {
      rAction.style.display='';
    } else {
      rAction.style.display='none';
    }
  }
}

async function doRevert(){
  if(!isAdmin) return;
  if(!confirm('Chắc chắn chạy lệnh REVERT phục hồi dựa trên log này?')) return;
  const r=await fetch('/api/revert',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({filename:curLogFile})}).then(x=>x.json()).catch(()=>null);
  if(r&&r.status==='ok') toast(r.msg, 'success');
  else toast((r&&r.error)||'Lỗi phục hồi!', 'error');
}

async function loadSettings(){
  if(!isAdmin) return;
  const cfg=await fetch('/api/config').then(r=>r.json()).catch(()=>({}));
  const m={'username':'cu','password':'cp','enable_password':'ce','vertiv_connect_password':'cv',
           'ssh_port':'csp','telnet_port':'ctp','menu_name_override':'cmno',
           'interval':'ci','verify_interval':'cvi','ip_list':'cil','baseline_db':'cbd','snapshot_db':'csd',
           'verify_wait_after_connect':'cvwac','max_verify_duration':'cmvd',
           'scan_schedule_time':'cst','scan_schedule_weekday':'csw','verify_schedule_time':'cvt','verify_schedule_weekday':'cvw'};
  for(const[k,id] of Object.entries(m)){const el=document.getElementById(id);if(el)el.value=cfg[k]??'';}
  const av=document.getElementById('cav');if(av)av.checked=cfg.auto_verify??true;
  const sm=document.getElementById('csm');if(sm){sm.value=cfg.scan_schedule_mode||'interval';togSF('scan');}
  const vm=document.getElementById('cvm');if(vm){vm.value=cfg.verify_schedule_mode||'interval';togSF('verify');}
}

function togSF(p){
  if(!isAdmin) return;
  const mode=document.getElementById(p==='scan'?'csm':'cvm').value;
  const tf=document.getElementById(p[0]+'_tf'),wf=document.getElementById(p[0]+'_wf');
  if(tf)tf.style.display=mode!=='interval'?'':'none';
  if(wf)wf.style.display=mode==='weekly'?'':'none';
}

async function saveCfg(){
  if(!isAdmin) return;
  const pay={username:g('cu').value,password:g('cp').value,enable_password:g('ce').value,
             vertiv_connect_password:g('cv').value,ssh_port:parseInt(g('csp').value)||22,
             telnet_port:parseInt(g('ctp').value)||23,menu_name_override:g('cmno').value,
             auto_verify:g('cav').checked};
  const r=await fetch('/api/config',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(pay)});
  if(r.ok)toast('Đã lưu cài đặt!','success');else toast('Lỗi lưu!','error');
}

async function saveSched(){
  if(!isAdmin) return;
  const pay={scan_schedule_mode:g('csm').value,interval:parseInt(g('ci').value)||30,
             scan_schedule_time:g('cst').value,scan_schedule_weekday:g('csw').value,
             verify_schedule_mode:g('cvm').value,verify_interval:parseInt(g('cvi').value)||3600,
             verify_schedule_time:g('cvt').value,verify_schedule_weekday:g('cvw').value,
             verify_wait_after_connect:parseFloat(g('cvwac').value)||1.5,max_verify_duration:parseInt(g('cmvd').value)||300};
  const r=await fetch('/api/config',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(pay)});
  if(r.ok)toast('Đã lưu lịch!','success');else toast('Lỗi!','error');
}

async function saveFiles(){
  if(!isAdmin) return;
  const pay={ip_list:g('cil').value,baseline_db:g('cbd').value,snapshot_db:g('csd').value};
  const r=await fetch('/api/config',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(pay)});
  if(r.ok)toast('Đã lưu!','success');else toast('Lỗi!','error');
}

async function loadCreds(){
  if(!isAdmin) return;
  const creds=await fetch('/api/credentials').then(r=>r.json()).catch(()=>[]);
  const el=document.getElementById('credList');
  if(!creds.length){el.innerHTML='<div style="color:var(--text3);font-size:13px;padding:12px">(Chưa có tài khoản phụ)</div>';return;}
  el.innerHTML=creds.map((c,i)=>`<div style="display:flex;align-items:center;gap:12px;padding:10px 14px;background:var(--bg-input);border:1px solid var(--border);border-radius:var(--rs);margin-bottom:8px"><span style="font-size:18px">👤</span><div style="flex:1"><div style="font-weight:600;font-size:13.5px">${esc(c.username)}</div><div style="font-size:11px;color:var(--text3)">${c.has_pass?'🔑 Có pass':''} ${c.has_enable?'· Enable ✓':''}</div></div><button class="btn btn-d btn-sm" onclick="delCred(${i})">Xóa</button></div>`).join('');
}

async function addCred(){
  if(!isAdmin) return;
  const user=g('cru').value.trim();
  if(!user){toast('Cần nhập username!','error');return;}
  await fetch('/api/credentials',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({username:user,password:g('crp').value,enable_password:g('cre').value})});
  cModal('addCred');toast('Đã thêm!','success');loadCreds();
}

async function delCred(idx){
  if(!isAdmin) return;
  if(!confirm('Xóa tài khoản này?'))return;
  await fetch('/api/credentials',{method:'DELETE',headers:{'Content-Type':'application/json'},body:JSON.stringify({index:idx})});
  toast('Đã xóa!','success');loadCreds();
}

async function doImport(){
  if(!isAdmin) return;
  const text=document.getElementById('importTxt').value.trim();
  if(!text){toast('Vui lòng nhập dữ liệu!','error');return;}
  const res=await fetch('/api/import',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({text})}).then(r=>r.json()).catch(()=>null);
  const el=document.getElementById('importRes');el.style.display='';
  if(res&&res.added!==undefined){el.innerHTML='<span class="text-t">✓ Thêm '+res.added+' thiết bị, bỏ qua '+res.skipped+'.</span>';toast('Import xong! +'+res.added,'success');}
  else{el.innerHTML='<span style="color:var(--red)">✗ Lỗi import!</span>';toast('Lỗi!','error');}
}

function oModal(id){
    if (!isAdmin && (id === 'addDev' || id === 'addCred' || id === 'settingsModal' || id === 'changePassMod')) {
        window.location.href='/login';
        return;
    }
    document.getElementById(id).classList.add('open');
}
function cModal(id){document.getElementById(id).classList.remove('open');}
document.addEventListener('click',e=>{if(e.target.classList.contains('mo'))e.target.classList.remove('open');});
document.addEventListener('keydown',e=>{if(e.key==='Escape')document.querySelectorAll('.mo.open').forEach(m=>m.classList.remove('open'));});

function toast(msg,type='info'){
  const c=document.getElementById('toastCnt');
  const t=document.createElement('div');t.className='toast '+type;
  const icons={success:'✅',error:'❌',info:'ℹ️',warning:'⚠️'};
  t.innerHTML='<span>'+(icons[type]||'ℹ️')+'</span><span style="flex:1">'+esc(msg)+'</span>';
  c.appendChild(t);setTimeout(()=>{t.style.opacity='0';t.style.transform='translateX(20px)';t.style.transition='.3s';setTimeout(()=>t.remove(),300);},4000);
}

function esc(s){return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');}
function nw(){return new Date().toLocaleTimeString('vi-VN');}
function g(id){return document.getElementById(id);}

setInterval(()=>{document.getElementById('clk').textContent=new Date().toLocaleString('vi-VN');},1000);
document.getElementById('clk').textContent=new Date().toLocaleString('vi-VN');

async function chkDaemon(){
  try{
    const tasks=await fetch('/api/tasks').then(r=>r.json());
    const run=Object.values(tasks).filter(t=>t.status==='running').length;
    const dot=document.getElementById('dDot'),txt=document.getElementById('dTxt');
    if(run>0){dot.className='d-dot on';txt.textContent=run+' task đang chạy';}
    else{dot.className='d-dot';txt.textContent='Web server hoạt động';}
  }catch{}
}
setInterval(chkDaemon,8000);chkDaemon();
initSSE();loadDash();
</script>
</body>
</html>"""

@app.route("/")
def index():
    is_admin = session.get('logged_in', False)
    return render_template_string(HTML, is_admin=is_admin)

@app.route("/device/<ip>")
def device_redir(ip):
    is_admin = session.get('logged_in', False)
    return render_template_string(HTML, is_admin=is_admin)

if __name__ == "__main__":
    print("=" * 60)
    print("  OOB Network Manager - Web Panel v2.3")
    print("  Dynamic UI: Phan quyen Guest (Read-only) / Admin (Action)")
    print("  Truy cap ngay: http://127.0.0.1:5000")
    print("=" * 60)
    app.run(host="0.0.0.0", port=5000, debug=False, threaded=True)